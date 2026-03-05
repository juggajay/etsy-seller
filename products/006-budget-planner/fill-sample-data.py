"""Fill sample data into the Navy budget planner so Dashboard charts render."""

import openpyxl
from datetime import date, timedelta

SRC = "/mnt/c/Users/jayso/businesses/etsy-templates/products/v2/output/006-budget-planner-v3/Budget-Planner-Navy.xlsx"
DST = "/mnt/c/Users/jayso/businesses/etsy-templates/products/v2/output/006-budget-planner-v3/Budget-Planner-Navy-SAMPLE.xlsx"

wb = openpyxl.load_workbook(SRC)

# ---- Monthly Budget: fill Budgeted column (B2:B16) ----
ws_budget = wb["Monthly Budget"]
budgets = [1800, 250, 600, 400, 200, 150, 500, 200, 300, 100, 50, 500, 80, 75, 150]
for i, amt in enumerate(budgets):
    ws_budget.cell(row=2 + i, column=2, value=amt)

# ---- Income Tracker: fill sample income ----
ws_income = wb["Income Tracker"]
income_data = [
    (date(2026, 3, 1),  "Salary",       "Monthly paycheck",       4500.00),
    (date(2026, 3, 5),  "Freelance",    "Logo design project",     800.00),
    (date(2026, 3, 10), "Dividends",    "Investment portfolio",    125.50),
    (date(2026, 3, 15), "Salary",       "Monthly paycheck",       4500.00),
    (date(2026, 3, 20), "Side Hustle",  "Etsy shop sales",         340.00),
]
for i, (dt, source, desc, amt) in enumerate(income_data):
    row = 2 + i
    ws_income.cell(row=row, column=1, value=dt)
    ws_income.cell(row=row, column=2, value=source)
    ws_income.cell(row=row, column=3, value=desc)
    ws_income.cell(row=row, column=4, value=amt)

# ---- Expense Tracker: fill sample expenses ----
ws_expense = wb["Expense Tracker"]
expenses = [
    (date(2026, 3, 1),  "Housing",             "Rent payment",               1800.00),
    (date(2026, 3, 1),  "Utilities",           "Electric bill",               120.00),
    (date(2026, 3, 2),  "Utilities",           "Internet service",             65.00),
    (date(2026, 3, 2),  "Subscriptions",       "Netflix + Spotify",            28.00),
    (date(2026, 3, 3),  "Food & Groceries",    "Weekly grocery shop",         145.00),
    (date(2026, 3, 4),  "Transportation",      "Gas fill-up",                  55.00),
    (date(2026, 3, 5),  "Insurance",           "Car insurance",               180.00),
    (date(2026, 3, 5),  "Healthcare",          "Prescription refill",          35.00),
    (date(2026, 3, 6),  "Entertainment",       "Movie tickets",                32.00),
    (date(2026, 3, 7),  "Food & Groceries",    "Coffee + lunch out",           24.50),
    (date(2026, 3, 8),  "Shopping",            "New running shoes",           120.00),
    (date(2026, 3, 9),  "Personal Care",       "Haircut",                      45.00),
    (date(2026, 3, 10), "Food & Groceries",    "Weekly grocery shop",         138.00),
    (date(2026, 3, 10), "Debt Payments",       "Student loan payment",        450.00),
    (date(2026, 3, 11), "Transportation",      "Uber rides",                   28.00),
    (date(2026, 3, 12), "Entertainment",       "Concert tickets",              85.00),
    (date(2026, 3, 13), "Food & Groceries",    "Takeout dinner",               42.00),
    (date(2026, 3, 14), "Gifts & Donations",   "Birthday gift for friend",     50.00),
    (date(2026, 3, 15), "Savings",             "Emergency fund deposit",      400.00),
    (date(2026, 3, 16), "Transportation",      "Gas fill-up",                  52.00),
    (date(2026, 3, 17), "Food & Groceries",    "Weekly grocery shop",         155.00),
    (date(2026, 3, 18), "Miscellaneous",       "Printer paper + ink",          65.00),
    (date(2026, 3, 19), "Shopping",            "Amazon order - kitchen",       78.00),
    (date(2026, 3, 20), "Education",           "Online course subscription",   29.99),
    (date(2026, 3, 21), "Subscriptions",       "iCloud storage",               2.99),
    (date(2026, 3, 22), "Food & Groceries",    "Weekend brunch",               38.00),
    (date(2026, 3, 23), "Entertainment",       "Video game purchase",          59.99),
    (date(2026, 3, 24), "Transportation",      "Car wash",                     15.00),
    (date(2026, 3, 25), "Healthcare",          "Dentist copay",                50.00),
    (date(2026, 3, 26), "Food & Groceries",    "Weekly grocery shop",         142.00),
    (date(2026, 3, 27), "Personal Care",       "Skincare products",            35.00),
    (date(2026, 3, 28), "Gifts & Donations",   "Charity donation",             25.00),
]
for i, (dt, cat, desc, amt) in enumerate(expenses):
    row = 2 + i
    ws_expense.cell(row=row, column=1, value=dt)
    ws_expense.cell(row=row, column=2, value=cat)
    ws_expense.cell(row=row, column=3, value=desc)
    ws_expense.cell(row=row, column=4, value=amt)

# ---- Savings Goals: fill sample goals ----
ws_goals = wb["Savings Goals"]
goals = [
    ("Emergency Fund",    10000, 4200, date(2026, 12, 31)),
    ("Vacation Fund",      3000, 1100, date(2026, 8, 15)),
    ("New Laptop",         2000,  650, date(2026, 10, 1)),
]
for i, (name, target, current, target_date) in enumerate(goals):
    row = 2 + i
    ws_goals.cell(row=row, column=1, value=name)
    ws_goals.cell(row=row, column=2, value=target)
    ws_goals.cell(row=row, column=3, value=current)
    ws_goals.cell(row=row, column=6, value=target_date)

wb.save(DST)
print(f"Saved sample data file to {DST}")
