"""
#006 Monthly Budget Planner Generator — follows xlsx-mastery skill exactly.
openpyxl. 4 colour themes. 6 tabs. Native charts.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from colours import THEMES

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- Constants ----------
CATEGORIES = [
    "Housing", "Utilities", "Food & Groceries", "Transportation",
    "Insurance", "Healthcare", "Debt Payments", "Entertainment",
    "Shopping", "Personal Care", "Education", "Savings",
    "Subscriptions", "Gifts & Donations", "Miscellaneous",
]
NUM_CATEGORIES = len(CATEGORIES)
INCOME_ROWS = 50
EXPENSE_ROWS = 200
GOAL_ROWS = 10

OUTPUT_DIR = "/mnt/c/Users/jayso/businesses/etsy-templates/products/v2/output/006-budget-planner-v3"

# ---------- Reusable styles (from xlsx-mastery skill) ----------

# Fonts
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10, color='333333')
DATA_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='333333')
TITLE_FONT_FN = lambda theme: Font(name='Arial', size=18, bold=True, color=theme["primary"])
TOTAL_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
# Blue text = hardcoded inputs users can change (financial model standard)
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')

# Fills
ZEBRA_FILL = PatternFill(fill_type='solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF')
GREEN_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')
RED_FILL = PatternFill(fill_type='solid', fgColor='FFC7CE')
YELLOW_FILL = PatternFill(fill_type='solid', fgColor='FFEB9C')

def header_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["primary"])

def accent_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["accent"])

def light_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["light_bg"])

# Borders
THIN_SIDE = Side(style='thin', color='D9D9D9')
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, bottom=THIN_SIDE, right=THIN_SIDE)

def header_border(theme):
    return Border(
        top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE,
        bottom=Side(style='medium', color=theme["primary"])
    )

TOTALS_BORDER = Border(
    top=Side(style='double'), bottom=Side(style='double'),
    left=THIN_SIDE, right=THIN_SIDE
)

# Alignment
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
LEFT_INDENT = Alignment(horizontal='left', vertical='center', indent=1)
RIGHT = Alignment(horizontal='right', vertical='center')


# ---------- Helpers ----------

def style_header_row(ws, row, start_col, end_col, headers, theme):
    """Style a header row per xlsx-mastery pattern."""
    for i, hdr in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = header_border(theme)
    ws.row_dimensions[row].height = 28


def style_data_row(ws, row, start_col, end_col, row_idx):
    """Apply thin borders + zebra stripe to all cells in a data row."""
    ws.row_dimensions[row].height = 22
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if row_idx % 2 == 0:
            cell.fill = ZEBRA_FILL
        else:
            cell.fill = WHITE_FILL


def setup_print(ws, num_cols, last_row, header_rows='1:1'):
    """Print setup per xlsx-mastery pattern."""
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_area = f'A1:{get_column_letter(num_cols)}{last_row}'
    ws.print_title_rows = header_rows


# ============================================================
# TAB 2: MONTHLY BUDGET (built first — Dashboard charts reference it)
# ============================================================

def build_monthly_budget(wb, theme):
    ws = wb.create_sheet("Monthly Budget")
    ws.sheet_properties.tabColor = theme["accent"]

    # Column widths
    for col, w in enumerate([25, 15, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row 1: Header
    headers = ["Category", "Budgeted", "Actual", "Difference", "Status"]
    style_header_row(ws, 1, 1, 5, headers, theme)

    # Freeze + auto filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(5)}1'

    # Data rows
    data_start = 2
    for i, cat in enumerate(CATEGORIES):
        row = data_start + i
        style_data_row(ws, row, 1, 5, i)

        # A: Category (locked)
        cell = ws.cell(row=row, column=1, value=cat)
        cell.font = DATA_FONT_BOLD
        cell.alignment = LEFT_INDENT

        # B: Budgeted (UNLOCKED — blue text = user input)
        cell = ws.cell(row=row, column=2, value=0)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # C: Actual = SUMIF from Expense Tracker
        exp_cat = f"'Expense Tracker'!B2:B{1 + EXPENSE_ROWS}"
        exp_amt = f"'Expense Tracker'!D2:D{1 + EXPENSE_ROWS}"
        cell = ws.cell(row=row, column=3)
        cell.value = f"=SUMIF({exp_cat},A{row},{exp_amt})"
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # D: Difference = Budgeted - Actual
        cell = ws.cell(row=row, column=4)
        cell.value = f'=B{row}-C{row}'
        cell.font = DATA_FONT_BOLD
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # E: Status
        cell = ws.cell(row=row, column=5)
        cell.value = f'=IF(D{row}>=0,"Under Budget","Over Budget")'
        cell.font = DATA_FONT

    # Totals row
    total_row = data_start + NUM_CATEGORIES
    ws.row_dimensions[total_row].height = 26
    ws.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = accent_fill(theme)
    ws.cell(row=total_row, column=1).alignment = LEFT_INDENT

    for col in [2, 3, 4]:
        letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM({letter}{data_start}:{letter}{total_row - 1})'
        cell.font = TOTAL_FONT
        cell.fill = accent_fill(theme)
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=5)
    cell.value = f'=IF(D{total_row}>=0,"Under Budget","Over Budget")'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.alignment = CENTER
    cell.border = TOTALS_BORDER

    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    # Conditional formatting: Difference column (green/red)
    diff_range = f'D{data_start}:D{total_row}'
    ws.conditional_formatting.add(
        diff_range,
        CellIsRule(operator='greaterThanOrEqual', formula=['0'],
                   fill=GREEN_FILL, font=Font(color='006100'))
    )
    ws.conditional_formatting.add(
        diff_range,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=RED_FILL, font=Font(color='9C0006'))
    )

    # Status column
    status_range = f'E{data_start}:E{total_row}'
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"Under Budget"'],
                   fill=GREEN_FILL, font=Font(color='006100', bold=True))
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"Over Budget"'],
                   fill=RED_FILL, font=Font(color='9C0006', bold=True))
    )

    setup_print(ws, 5, total_row)

    # Protection (Budgeted column unlocked via cell-level)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 3: INCOME TRACKER
# ============================================================

def build_income_tracker(wb, theme):
    ws = wb.create_sheet("Income Tracker")
    ws.sheet_properties.tabColor = theme["highlight"]

    for col, w in enumerate([14, 25, 32, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row 1: Header
    headers = ["Date", "Source", "Description", "Amount"]
    style_header_row(ws, 1, 1, 4, headers, theme)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:D1'

    # Data rows
    for i in range(INCOME_ROWS):
        row = 2 + i
        style_data_row(ws, row, 1, 4, i)

        ws.cell(row=row, column=1).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=1).protection = Protection(locked=False)
        ws.cell(row=row, column=1).font = INPUT_FONT

        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).protection = Protection(locked=False)
        ws.cell(row=row, column=2).font = INPUT_FONT

        ws.cell(row=row, column=3).alignment = LEFT
        ws.cell(row=row, column=3).protection = Protection(locked=False)
        ws.cell(row=row, column=3).font = INPUT_FONT

        ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4).alignment = RIGHT
        ws.cell(row=row, column=4).protection = Protection(locked=False)
        ws.cell(row=row, column=4).font = INPUT_FONT

    # Totals row
    total_row = 2 + INCOME_ROWS
    ws.row_dimensions[total_row].height = 26
    ws.cell(row=total_row, column=3, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=3).fill = accent_fill(theme)
    ws.cell(row=total_row, column=3).alignment = RIGHT

    cell = ws.cell(row=total_row, column=4)
    cell.value = f'=SUM(D2:D{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    for col in [1, 2, 3]:
        ws.cell(row=total_row, column=col).border = TOTALS_BORDER
        ws.cell(row=total_row, column=col).fill = accent_fill(theme)

    setup_print(ws, 4, total_row)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 4: EXPENSE TRACKER
# ============================================================

def build_expense_tracker(wb, theme):
    ws = wb.create_sheet("Expense Tracker")
    ws.sheet_properties.tabColor = theme["highlight"]

    for col, w in enumerate([14, 25, 32, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row 1: Header
    headers = ["Date", "Category", "Description", "Amount"]
    style_header_row(ws, 1, 1, 4, headers, theme)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:D1'

    # Category dropdown
    dv = DataValidation(
        type='list',
        formula1='"' + ','.join(CATEGORIES) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='Invalid Category',
        error='Please select from the dropdown list'
    )
    ws.add_data_validation(dv)

    # Data rows
    for i in range(EXPENSE_ROWS):
        row = 2 + i
        style_data_row(ws, row, 1, 4, i)

        ws.cell(row=row, column=1).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=1).protection = Protection(locked=False)
        ws.cell(row=row, column=1).font = INPUT_FONT

        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).protection = Protection(locked=False)
        ws.cell(row=row, column=2).font = INPUT_FONT

        ws.cell(row=row, column=3).alignment = LEFT
        ws.cell(row=row, column=3).protection = Protection(locked=False)
        ws.cell(row=row, column=3).font = INPUT_FONT

        ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4).alignment = RIGHT
        ws.cell(row=row, column=4).protection = Protection(locked=False)
        ws.cell(row=row, column=4).font = INPUT_FONT

    dv.add(f'B2:B{1 + EXPENSE_ROWS}')

    # Totals row
    total_row = 2 + EXPENSE_ROWS
    ws.row_dimensions[total_row].height = 26
    ws.cell(row=total_row, column=3, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=3).fill = accent_fill(theme)
    ws.cell(row=total_row, column=3).alignment = RIGHT

    cell = ws.cell(row=total_row, column=4)
    cell.value = f'=SUM(D2:D{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    for col in [1, 2, 3]:
        ws.cell(row=total_row, column=col).border = TOTALS_BORDER
        ws.cell(row=total_row, column=col).fill = accent_fill(theme)

    setup_print(ws, 4, total_row)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 5: SAVINGS GOALS
# ============================================================

def build_savings_goals(wb, theme):
    ws = wb.create_sheet("Savings Goals")
    ws.sheet_properties.tabColor = theme["accent"]

    for col, w in enumerate([25, 16, 16, 13, 20, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    headers = ["Goal Name", "Target Amount", "Current Saved",
               "Progress", "Monthly Contribution", "Target Date"]
    style_header_row(ws, 1, 1, 6, headers, theme)

    ws.freeze_panes = 'A2'

    data_start = 2
    for i in range(GOAL_ROWS):
        row = data_start + i
        style_data_row(ws, row, 1, 6, i)

        # Goal Name (unlocked, blue = input)
        cell = ws.cell(row=row, column=1)
        cell.font = Font(name='Arial', size=10, bold=True, color='0000FF')
        cell.alignment = LEFT_INDENT
        cell.protection = Protection(locked=False)

        # Target Amount (unlocked, blue)
        cell = ws.cell(row=row, column=2)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # Current Saved (unlocked, blue)
        cell = ws.cell(row=row, column=3)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # Progress % (locked — formula, black text)
        cell = ws.cell(row=row, column=4)
        cell.value = f'=IF(B{row}=0,"",C{row}/B{row})'
        cell.font = Font(name='Arial', size=10, bold=True, color='000000')
        cell.number_format = '0%'

        # Monthly Contribution (locked — formula, black text)
        cell = ws.cell(row=row, column=5)
        cell.value = (
            f'=IF(OR(B{row}=0,F{row}=""),"",IF(F{row}<=TODAY(),0,'
            f'(B{row}-C{row})/MAX(1,(F{row}-TODAY())/30.44)))'
        )
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # Target Date (unlocked, blue)
        cell = ws.cell(row=row, column=6)
        cell.font = INPUT_FONT
        cell.number_format = 'mm/dd/yyyy'
        cell.protection = Protection(locked=False)

    # Data bars on Progress column
    ws.conditional_formatting.add(
        f'D{data_start}:D{data_start + GOAL_ROWS - 1}',
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=1,
                    color=theme["highlight"])
    )

    setup_print(ws, 6, data_start + GOAL_ROWS - 1)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 6: INSTRUCTIONS
# ============================================================

def build_instructions(wb, theme):
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = theme["primary"]
    ws.column_dimensions['A'].width = 85

    lines = [
        ("How to Use Your Budget Planner", True, 18, theme["primary"]),
        ("", False, 11, None),
        ("Welcome to your FileSmartCo Monthly Budget Planner!", False, 11, None),
        ("This spreadsheet is designed to make managing your money simple.", False, 11, None),
        ("", False, 11, None),
        ("GETTING STARTED", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("1.  Start with the Expense Tracker tab \u2014 your primary data entry point.", False, 11, None),
        ("     Enter each expense: date, category (dropdown), description, amount.", False, 10, '808080'),
        ("     Blue text = cells you can edit. Black text = auto-calculated formulas.", False, 10, '808080'),
        ("", False, 6, None),
        ("2.  Go to the Income Tracker and enter your income sources.", False, 11, None),
        ("     Add each payment: date, source, description, amount.", False, 10, '808080'),
        ("", False, 6, None),
        ("3.  Set your budget in the Monthly Budget tab.", False, 11, None),
        ("     Enter planned amounts in the Budgeted column (blue text).", False, 10, '808080'),
        ("     The Actual column auto-calculates from your Expense Tracker.", False, 10, '808080'),
        ("", False, 6, None),
        ("4.  Check the Dashboard for your overview with charts.", False, 11, None),
        ("     Everything updates automatically \u2014 no manual math!", False, 10, '808080'),
        ("", False, 6, None),
        ("5.  Set savings goals in the Savings Goals tab.", False, 11, None),
        ("     Progress and monthly contribution are calculated for you.", False, 10, '808080'),
        ("", False, 11, None),
        ("CUSTOMISING CATEGORIES", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("\u2022  Unprotect the Monthly Budget sheet (Review > Unprotect Sheet)", False, 11, None),
        ("\u2022  Edit category names in column A", False, 11, None),
        ("\u2022  Update the dropdown in Expense Tracker if needed", False, 11, None),
        ("\u2022  Re-protect the sheet when done", False, 11, None),
        ("", False, 11, None),
        ("ADDING MORE ROWS", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("Income Tracker: 50 rows. Expense Tracker: 200 rows.", False, 11, None),
        ("To add more: unprotect, copy last row, paste below, extend total formula.", False, 11, None),
        ("", False, 11, None),
        ("BUDGETING TIPS \u2014 The 50/30/20 Rule", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("     50% \u2192 Needs (housing, utilities, food, transport)", False, 11, None),
        ("     30% \u2192 Wants (entertainment, shopping, subscriptions)", False, 11, None),
        ("     20% \u2192 Savings & Debt Payments", False, 11, None),
        ("", False, 6, None),
        ("\u2022  Track expenses daily \u2014 don't let them pile up", False, 11, None),
        ("\u2022  Review your dashboard weekly", False, 11, None),
        ("\u2022  Adjust your budget monthly based on actual spending", False, 11, None),
        ("\u2022  Set realistic goals \u2014 start small and build up", False, 11, None),
        ("", False, 11, None),
        ("FOR GOOGLE SHEETS USERS", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("1.  Open the link from your download", False, 11, None),
        ("2.  File > Make a copy", False, 11, None),
        ("3.  Save to your Google Drive", False, 11, None),
        ("Do NOT edit the original \u2014 always make a copy first.", True, 11, '9C0006'),
        ("", False, 11, None),
        ("VIDEO TUTORIAL", True, 13, theme["primary"]),
        ("Watch our video tutorial: [LINK COMING SOON]", False, 11, '808080'),
        ("", False, 11, None),
        ("SUPPORT", True, 13, theme["primary"]),
        ("Questions? Message us on Etsy \u2014 we're happy to help!", False, 11, None),
        ("", False, 11, None),
        ("Thank you for choosing FileSmartCo!", True, 13, theme["primary"]),
    ]

    for i, (text, bold, size, color) in enumerate(lines):
        cell = ws.cell(row=i + 1, column=1, value=text)
        cell.font = Font(name='Arial', bold=bold, size=size,
                         color=color or '333333')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 1: DASHBOARD (built last — references other tabs)
# ============================================================

def build_dashboard(wb, theme):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = theme["primary"]

    # FIX #1: Wider columns so dollar amounts never show ########
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Row 1: Title
    ws.merge_cells('A1:H1')
    ws['A1'].value = 'Monthly Budget Dashboard'
    ws['A1'].font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # Row 2: spacer
    ws.row_dimensions[2].height = 6

    # FIX #6: Cards span 2 cols each, flush against each other (no gap columns)
    # Layout: A-B = Income, C-D = Expenses, E-F = Net Savings, G-H = Savings Rate
    ws.row_dimensions[3].height = 26

    primary_border = Side(style='thin', color=theme["primary"])
    card_header_border = Border(top=primary_border, left=primary_border,
                                right=primary_border, bottom=primary_border)

    card_spans = [('A3:B3', 1, 'TOTAL INCOME'), ('C3:D3', 3, 'TOTAL EXPENSES'),
                  ('E3:F3', 5, 'NET SAVINGS'), ('G3:H3', 7, 'SAVINGS RATE')]
    for merge_range, col, label in card_spans:
        ws.merge_cells(merge_range)
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = card_header_border
        # Fill + border on second cell of merge (Excel needs this for visual continuity)
        c2 = ws.cell(row=3, column=col + 1)
        c2.fill = header_fill(theme)
        c2.border = card_header_border

    # Row 4: Card values (merged spans, flush)
    ws.row_dimensions[4].height = 40
    link_font = Font(name='Arial', size=16, bold=True, color='008000')
    calc_font = Font(name='Arial', size=16, bold=True, color='000000')

    card_val_border = Border(
        left=primary_border, right=primary_border,
        bottom=Side(style='medium', color=theme["accent"])
    )

    value_merges = ['A4:B4', 'C4:D4', 'E4:F4', 'G4:H4']
    for merge_range in value_merges:
        ws.merge_cells(merge_range)

    # Style all value cells with fill + border (both cells in each merge)
    for col in range(1, 9):
        c = ws.cell(row=4, column=col)
        c.fill = light_fill(theme)
        c.border = card_val_border

    # A4: Total Income
    cell = ws.cell(row=4, column=1)
    cell.value = f"='Income Tracker'!D{2 + INCOME_ROWS}"
    cell.font = link_font
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # C4: Total Expenses
    cell = ws.cell(row=4, column=3)
    cell.value = f"='Expense Tracker'!D{2 + EXPENSE_ROWS}"
    cell.font = link_font
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # E4: Net Savings
    cell = ws.cell(row=4, column=5)
    cell.value = '=A4-C4'
    cell.font = calc_font
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # G4: Savings Rate — FIX #9: show blank when no income (neutral, no red 0%)
    cell = ws.cell(row=4, column=7)
    cell.value = '=IF(A4=0,"",E4/A4)'
    cell.font = calc_font
    cell.number_format = '0.0%'
    cell.alignment = CENTER

    # FIX #9: Only trigger conditional formatting when Total Income > 0
    ws.conditional_formatting.add('G4',
        FormulaRule(formula=['AND(A4>0,G4>=0.2)'],
                    fill=GREEN_FILL, font=Font(color='006100', bold=True, size=16)))
    ws.conditional_formatting.add('G4',
        FormulaRule(formula=['AND(A4>0,G4>=0.1,G4<0.2)'],
                    fill=YELLOW_FILL, font=Font(color='9C5700', bold=True, size=16)))
    ws.conditional_formatting.add('G4',
        FormulaRule(formula=['AND(A4>0,G4<0.1)'],
                    fill=RED_FILL, font=Font(color='9C0006', bold=True, size=16)))

    # Row 5: spacer
    ws.row_dimensions[5].height = 6

    # Row 6: Section header
    ws.merge_cells('A6:H6')
    ws['A6'].value = 'Spending Overview'
    ws['A6'].font = Font(name='Arial', size=13, bold=True, color=theme["primary"])
    ws['A6'].alignment = LEFT
    ws.row_dimensions[6].height = 26
    for col in range(1, 9):
        ws.cell(row=6, column=col).border = Border(
            bottom=Side(style='medium', color=theme["accent"]))

    # --- PIE CHART ---
    pie = PieChart()
    pie.title = 'Spending by Category'
    pie.style = 10

    # Reference Actual column directly (no helper row)
    cats = Reference(wb["Monthly Budget"], min_col=1, min_row=2, max_row=1 + NUM_CATEGORIES)
    vals = Reference(wb["Monthly Budget"], min_col=3, min_row=2, max_row=1 + NUM_CATEGORIES)
    pie.add_data(vals, titles_from_data=False)
    pie.set_categories(cats)
    pie.width = 18
    pie.height = 13

    palette = theme["chart_palette"]
    slices = []
    for i in range(NUM_CATEGORIES):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = palette[i % len(palette)]
        slices.append(dp)
    pie.series[0].data_points = slices

    # FIX #5: NO data labels on pie chart — legend provides category names
    # This prevents the broken cluttered "0%" labels when empty
    # When populated, slice sizes + legend colors convey the information
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = False
    pie.dataLabels.showCatName = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showVal = False
    pie.legend.position = 'r'
    pie.legend.overlay = False

    ws.add_chart(pie, 'A7')

    # --- BAR CHART ---
    bar = BarChart()
    bar.type = 'col'
    bar.grouping = 'clustered'
    bar.title = 'Budget vs Actual'
    bar.style = 10
    bar.y_axis.title = 'Amount ($)'
    bar.y_axis.numFmt = '"$"#,##0'

    # FIX #3: Single add_data call with both columns (B:C) — standard openpyxl pattern
    # Row 1 has headers "Budgeted" and "Actual", titles_from_data reads them
    data_ref = Reference(wb["Monthly Budget"], min_col=2, max_col=3,
                         min_row=1, max_row=1 + NUM_CATEGORIES)
    bar_cats = Reference(wb["Monthly Budget"], min_col=1, min_row=2, max_row=1 + NUM_CATEGORIES)

    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.x_axis.tickLblPos = 'low'
    bar.x_axis.delete = False

    bar.series[0].graphicalProperties.solidFill = theme["chart_budget"]
    bar.series[1].graphicalProperties.solidFill = theme["chart_actual"]

    bar.width = 22
    bar.height = 13
    bar.legend.position = 'b'
    bar.legend.overlay = False

    ws.add_chart(bar, 'A23')

    # Dashboard is fully protected (no user input)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# MAIN
# ============================================================

def generate_workbook(theme_name, theme):
    wb = Workbook()
    temp = wb.active
    temp.title = '_temp'

    # Build in dependency order
    build_monthly_budget(wb, theme)
    build_income_tracker(wb, theme)
    build_expense_tracker(wb, theme)
    build_savings_goals(wb, theme)
    build_instructions(wb, theme)
    build_dashboard(wb, theme)  # last — charts reference other tabs

    del wb['_temp']

    wb.properties.creator = 'FileSmartCo'
    wb.properties.title = f'Monthly Budget Planner - {theme_name}'
    wb.properties.company = 'FileSmartCo'

    return wb


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for name, theme in THEMES.items():
        fname = f'Budget-Planner-{name}.xlsx'
        path = os.path.join(OUTPUT_DIR, fname)
        print(f'Generating {fname}...')
        wb = generate_workbook(name, theme)
        wb.save(path)
        print(f'  \u2713 {path}')
    print('\nDone! All 4 variants generated.')


if __name__ == '__main__':
    main()
