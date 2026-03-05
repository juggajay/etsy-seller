"""
#007 Debt Payoff Tracker Generator — follows xlsx-mastery skill exactly.
openpyxl. 4 colour themes. 6 tabs + 2 hidden calc sheets. Native charts.
Same pattern as #006 Budget Planner.

V2: Proper month-by-month amortization for snowball/avalanche cascade.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from colours import THEMES

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- Constants ----------
DEBT_ROWS = 15
PAYMENT_LOG_ROWS = 200
MAX_MONTHS = 120  # 10-year cap on amortization grid

DEBT_TYPES = [
    "Credit Card", "Student Loan", "Car Loan", "Mortgage",
    "Personal Loan", "Medical", "Other",
]

SAMPLE_DEBTS = [
    ("Credit Card", "Credit Card", 4500, 22.9, 90),
    ("Student Loan", "Student Loan", 28000, 5.5, 295),
    ("Car Loan", "Car Loan", 12000, 6.9, 350),
    ("Personal Loan", "Personal Loan", 3200, 11.0, 85),
    ("Medical Debt", "Medical", 1800, 0.0, 75),
]

OUTPUT_DIR = "/mnt/c/Users/jayso/businesses/etsy-templates/products/v2/output/007-debt-payoff-tracker"

# ---------- Reusable styles (from xlsx-mastery / #006 pattern) ----------

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10, color='333333')
DATA_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='333333')
TITLE_FONT_FN = lambda theme: Font(name='Arial', size=18, bold=True, color=theme["primary"])
TOTAL_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')
INPUT_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='0000FF')

ZEBRA_FILL = PatternFill(fill_type='solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF')
GREEN_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')
RED_FILL = PatternFill(fill_type='solid', fgColor='FFC7CE')
YELLOW_FILL = PatternFill(fill_type='solid', fgColor='FFEB9C')


def header_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["primary"])

def accent_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["accent"])

def light_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["light_bg"])

THIN_SIDE = Side(style='thin', color='D9D9D9')
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, bottom=THIN_SIDE, right=THIN_SIDE)
TOTALS_BORDER = Border(top=Side(style='double'), bottom=Side(style='double'),
                        left=THIN_SIDE, right=THIN_SIDE)

def header_border(theme):
    return Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE,
                  bottom=Side(style='medium', color=theme["primary"]))

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
LEFT_INDENT = Alignment(horizontal='left', vertical='center', indent=1)
RIGHT = Alignment(horizontal='right', vertical='center')


# ---------- Helpers ----------

def style_header_row(ws, row, start_col, end_col, headers, theme):
    for i, hdr in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = header_border(theme)
    ws.row_dimensions[row].height = 28

def style_data_row(ws, row, start_col, end_col, row_idx):
    ws.row_dimensions[row].height = 22
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        if row_idx % 2 == 0:
            cell.fill = ZEBRA_FILL
        else:
            cell.fill = WHITE_FILL

def setup_print(ws, num_cols, last_row, header_rows='1:1'):
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_area = f'A1:{get_column_letter(num_cols)}{last_row}'
    ws.print_title_rows = header_rows


# ============================================================
# TAB 1: DEBT INVENTORY (unchanged from v1)
# ============================================================

def build_debt_inventory(wb, theme):
    ws = wb.create_sheet("Debt Inventory")
    ws.sheet_properties.tabColor = theme["accent"]

    for col, w in enumerate([22, 18, 16, 14, 16, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    headers = ["Debt Name", "Type", "Balance", "Interest Rate", "Min Payment", "Status"]
    style_header_row(ws, 1, 1, 6, headers, theme)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:F1'

    dv_type = DataValidation(
        type='list',
        formula1='"' + ','.join(DEBT_TYPES) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Type', error='Please select a debt type from the dropdown'
    )
    ws.add_data_validation(dv_type)

    data_start = 2
    for i in range(DEBT_ROWS):
        row = data_start + i
        style_data_row(ws, row, 1, 6, i)

        ws.cell(row=row, column=1).font = INPUT_FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_INDENT
        ws.cell(row=row, column=1).protection = Protection(locked=False)

        ws.cell(row=row, column=2).font = INPUT_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).protection = Protection(locked=False)

        cell = ws.cell(row=row, column=3)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        cell = ws.cell(row=row, column=4)
        cell.font = INPUT_FONT
        cell.number_format = '0.0%'
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)

        cell = ws.cell(row=row, column=5)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        cell = ws.cell(row=row, column=6)
        cell.value = f'=IF(C{row}="","",IF(C{row}<=0,"Paid Off","Active"))'
        cell.font = DATA_FONT_BOLD
        cell.alignment = CENTER

        if i < len(SAMPLE_DEBTS):
            name, dtype, balance, rate, minpay = SAMPLE_DEBTS[i]
            ws.cell(row=row, column=1).value = name
            ws.cell(row=row, column=2).value = dtype
            ws.cell(row=row, column=3).value = balance
            ws.cell(row=row, column=4).value = rate / 100
            ws.cell(row=row, column=5).value = minpay

    dv_type.add(f'B{data_start}:B{data_start + DEBT_ROWS - 1}')

    total_row = data_start + DEBT_ROWS
    ws.row_dimensions[total_row].height = 26
    ws.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = accent_fill(theme)
    ws.cell(row=total_row, column=1).alignment = LEFT_INDENT
    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    for col in [2, 4, 6]:
        c = ws.cell(row=total_row, column=col)
        c.fill = accent_fill(theme)
        c.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=3)
    cell.value = f'=SUM(C{data_start}:C{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=5)
    cell.value = f'=SUM(E{data_start}:E{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    count_row = total_row + 1
    ws.cell(row=count_row, column=1, value='Active Debts').font = DATA_FONT
    cell = ws.cell(row=count_row, column=3)
    cell.value = f'=COUNTIF(F{data_start}:F{total_row - 1},"Active")'
    cell.font = DATA_FONT_BOLD

    status_range = f'F{data_start}:F{total_row - 1}'
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Active"'],
                   fill=YELLOW_FILL, font=Font(color='9C5700', bold=True)))
    ws.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Paid Off"'],
                   fill=GREEN_FILL, font=Font(color='006100', bold=True)))

    setup_print(ws, 6, total_row)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# HIDDEN: AMORTIZATION CALC SHEET (used by both methods)
# ============================================================

def build_method_calc(wb, method):
    """Build hidden month-by-month amortization sheet.

    method: 'snowball' or 'avalanche'

    Layout:
      Row 1: Starting balances (from sorted plan tab C6:C20)
      Row 2: Interest rates (from sorted plan tab D6:D20)
      Row 3: Min payments (from sorted plan tab E6:E20)
      Row 4: Total budget (from plan tab D3)
      Row 5: (spacer)
      Row 6: Month 0 — starting balances
      Rows 7-126: Months 1-120 — amortization formulas
      Columns B-P: Balance for sorted debts 1-15
      Column Q: Monthly interest
      Column R: Cumulative interest
    """
    if method == 'snowball':
        sheet_name = "Snowball Calc"
        plan_tab = "'Snowball Plan'"
    else:
        sheet_name = "Avalanche Calc"
        plan_tab = "'Avalanche Plan'"

    ws = wb.create_sheet(sheet_name)
    ws.sheet_state = 'hidden'

    last_debt_col = DEBT_ROWS + 1  # column 16 = P
    last_debt_letter = get_column_letter(last_debt_col)
    int_col = DEBT_ROWS + 2   # column Q
    cum_col = DEBT_ROWS + 3   # column R
    int_letter = get_column_letter(int_col)
    cum_letter = get_column_letter(cum_col)

    # Row 1-3: Reference data from sorted plan tab (coerce "" to 0)
    for j in range(DEBT_ROWS):
        col = j + 2
        plan_row = 6 + j
        ws.cell(row=1, column=col).value = f'=IF({plan_tab}!C{plan_row}="",0,{plan_tab}!C{plan_row})'
        ws.cell(row=2, column=col).value = f'=IF({plan_tab}!D{plan_row}="",0,{plan_tab}!D{plan_row})'
        ws.cell(row=3, column=col).value = f'=IF({plan_tab}!E{plan_row}="",0,{plan_tab}!E{plan_row})'

    # Row 4: Total budget
    ws.cell(row=4, column=2).value = f'={plan_tab}!D3'

    # Row 6: Month 0 — starting balances
    ws.cell(row=6, column=1, value=0)
    for j in range(DEBT_ROWS):
        col = j + 2
        cl = get_column_letter(col)
        ws.cell(row=6, column=col).value = f'={cl}1'
    ws.cell(row=6, column=int_col, value=0)
    ws.cell(row=6, column=cum_col, value=0)

    # Rows 7-126: Months 1-120
    for m in range(1, MAX_MONTHS + 1):
        row = 6 + m
        prev_row = row - 1
        ws.cell(row=row, column=1, value=m)

        # Ranges for SUMPRODUCT
        bal_range = f'B{prev_row}:{last_debt_letter}{prev_row}'
        min_range = f'B$3:{last_debt_letter}$3'
        rate_range = f'B$2:{last_debt_letter}$2'

        # Extra available = TotalBudget - SUM(min payments of active debts)
        extra_avail = f'$B$4-SUMPRODUCT({min_range},--({bal_range}>0))'

        for j in range(DEBT_ROWS):
            col = j + 2
            cl = get_column_letter(col)
            prev_bal = f'{cl}{prev_row}'
            rate = f'{cl}$2'
            min_pay = f'{cl}$3'

            # isPriority: this debt has balance > 0 AND all debts before it
            # (in sorted order) have balance <= 0
            if j == 0:
                is_priority = f'{prev_bal}>0'
            else:
                conds = ','.join(
                    f'{get_column_letter(k + 2)}{prev_row}<=0' for k in range(j)
                )
                is_priority = f'AND({prev_bal}>0,{conds})'

            # New balance = MAX(0, PrevBal + Interest - MinPay - ExtraIfPriority)
            formula = (
                f'=IF({prev_bal}<=0,0,'
                f'MAX(0,{prev_bal}+{prev_bal}*{rate}/12-{min_pay}'
                f'-IF({is_priority},{extra_avail},0)))'
            )
            ws.cell(row=row, column=col).value = formula

        # Monthly interest = SUM of (PrevBal * Rate / 12) for active debts
        ws.cell(row=row, column=int_col).value = (
            f'=SUMPRODUCT(--({bal_range}>0),{bal_range},{rate_range})/12'
        )

        # Cumulative interest
        ws.cell(row=row, column=cum_col).value = (
            f'={cum_letter}{prev_row}+{int_letter}{row}'
        )

    return ws


# ============================================================
# TAB 2: SNOWBALL PLAN (sort by balance ascending)
# ============================================================

def build_snowball_plan(wb, theme):
    ws = wb.create_sheet("Snowball Plan")
    ws.sheet_properties.tabColor = theme["highlight"]

    for col, w in enumerate([6, 22, 16, 14, 16, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    inv = "'Debt Inventory'"
    calc = "'Snowball Calc'"
    data_start_inv = 2
    data_end_inv = data_start_inv + DEBT_ROWS - 1
    last_calc_row = 6 + MAX_MONTHS  # row 126

    # Row 1: Title
    ws.merge_cells('A1:H1')
    ws['A1'].value = 'Debt Snowball Plan \u2014 Smallest Balance First'
    ws['A1'].font = TITLE_FONT_FN(theme)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    # Row 2: Extra monthly payment
    ws.cell(row=2, column=1, value='Extra Monthly Payment:').font = DATA_FONT_BOLD
    ws.merge_cells('A2:C2')
    cell = ws.cell(row=2, column=4, value=200)
    cell.font = INPUT_FONT_BOLD
    cell.number_format = '"$"#,##0.00'
    cell.protection = Protection(locked=False)
    cell.alignment = RIGHT
    ws.row_dimensions[2].height = 28

    # Row 3: Total monthly payment
    ws.cell(row=3, column=1, value='Total Monthly Payment:').font = DATA_FONT_BOLD
    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=4)
    cell.value = f"={inv}!E{data_start_inv + DEBT_ROWS}+D2"
    cell.font = DATA_FONT_BOLD
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    ws.row_dimensions[3].height = 24

    ws.row_dimensions[4].height = 6

    # Row 5: Headers (no "Total Payment" column — meaningless with cascade)
    headers = ["#", "Debt Name", "Balance", "Rate", "Min Payment",
               "Months", "Interest Paid", "Payoff Date"]
    style_header_row(ws, 5, 1, 8, headers, theme)
    ws.freeze_panes = 'A6'

    summary_start = 6
    for i in range(DEBT_ROWS):
        row = summary_start + i
        style_data_row(ws, row, 1, 8, i)
        rank = i + 1
        calc_col = get_column_letter(i + 2)  # B, C, D, ... P in calc sheet

        # A: Priority #
        ws.cell(row=row, column=1, value=rank).font = DATA_FONT_BOLD
        ws.cell(row=row, column=1).alignment = CENTER

        # B: Debt Name (sorted by SMALL balance)
        cell = ws.cell(row=row, column=2)
        cell.value = (
            f'=IFERROR(INDEX({inv}!A{data_start_inv}:A{data_end_inv},'
            f'MATCH(SMALL({inv}!C{data_start_inv}:C{data_end_inv},{rank}),'
            f'{inv}!C{data_start_inv}:C{data_end_inv},0)),"")'
        )
        cell.font = DATA_FONT_BOLD
        cell.alignment = LEFT_INDENT

        # C: Balance
        cell = ws.cell(row=row, column=3)
        cell.value = f'=IFERROR(SMALL({inv}!C{data_start_inv}:C{data_end_inv},{rank}),"")'
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # D: Interest Rate
        cell = ws.cell(row=row, column=4)
        cell.value = (
            f'=IFERROR(INDEX({inv}!D{data_start_inv}:D{data_end_inv},'
            f'MATCH(SMALL({inv}!C{data_start_inv}:C{data_end_inv},{rank}),'
            f'{inv}!C{data_start_inv}:C{data_end_inv},0)),"")'
        )
        cell.font = DATA_FONT
        cell.number_format = '0.0%'
        cell.alignment = CENTER

        # E: Min Payment
        cell = ws.cell(row=row, column=5)
        cell.value = (
            f'=IFERROR(INDEX({inv}!E{data_start_inv}:E{data_end_inv},'
            f'MATCH(SMALL({inv}!C{data_start_inv}:C{data_end_inv},{rank}),'
            f'{inv}!C{data_start_inv}:C{data_end_inv},0)),"")'
        )
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # F: Months to payoff (from calc sheet)
        # COUNTIF counts months where balance > 0 = number of payments needed
        cell = ws.cell(row=row, column=6)
        cell.value = (
            f'=IF(OR(C{row}="",C{row}=0),"",'
            f'IF({calc}!{calc_col}{last_calc_row}>0,"N/A",'
            f'COUNTIF({calc}!{calc_col}$6:{calc_col}${last_calc_row},">0")))'
        )
        cell.font = DATA_FONT_BOLD
        cell.alignment = CENTER

        # G: Interest Paid = Rate/12 * SUM(balances month 0 to month 119)
        # This is exact: interest each month = prev_balance * rate/12
        cell = ws.cell(row=row, column=7)
        cell.value = (
            f'=IF(OR(C{row}="",C{row}=0),"",'
            f'D{row}/12*SUM({calc}!{calc_col}$6:{calc_col}${last_calc_row - 1}))'
        )
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # H: Payoff Date
        cell = ws.cell(row=row, column=8)
        cell.value = (
            f'=IF(OR(F{row}="",F{row}="N/A"),"",'
            f'DATE(YEAR(TODAY()),MONTH(TODAY())+F{row},DAY(TODAY())))'
        )
        cell.font = DATA_FONT
        cell.number_format = 'mmm yyyy'
        cell.alignment = CENTER

    # Totals row
    summary_end = summary_start + DEBT_ROWS - 1
    total_row = summary_end + 1
    ws.row_dimensions[total_row].height = 26

    ws.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = accent_fill(theme)
    ws.cell(row=total_row, column=1).alignment = LEFT_INDENT
    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    for col in [2, 4, 5, 6, 8]:
        c = ws.cell(row=total_row, column=col)
        c.fill = accent_fill(theme)
        c.border = TOTALS_BORDER

    # Total Balance
    cell = ws.cell(row=total_row, column=3)
    cell.value = f'=SUM(C{summary_start}:C{summary_end})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Total Interest (from calc sheet cumulative)
    cell = ws.cell(row=total_row, column=7)
    cell.value = f'={calc}!{get_column_letter(DEBT_ROWS + 3)}{last_calc_row}'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Conditional formatting: months column
    months_range = f'F{summary_start}:F{summary_end}'
    ws.conditional_formatting.add(months_range,
        CellIsRule(operator='equal', formula=['"N/A"'],
                   fill=RED_FILL, font=Font(color='9C0006', bold=True)))

    setup_print(ws, 8, total_row, '5:5')
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 3: AVALANCHE PLAN (sort by interest rate descending)
# ============================================================

def build_avalanche_plan(wb, theme):
    ws = wb.create_sheet("Avalanche Plan")
    ws.sheet_properties.tabColor = theme["highlight"]

    for col, w in enumerate([6, 22, 16, 14, 16, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    inv = "'Debt Inventory'"
    calc = "'Avalanche Calc'"
    snow_calc = "'Snowball Calc'"
    data_start_inv = 2
    data_end_inv = data_start_inv + DEBT_ROWS - 1
    last_calc_row = 6 + MAX_MONTHS
    cum_col_letter = get_column_letter(DEBT_ROWS + 3)  # R

    # Row 1: Title
    ws.merge_cells('A1:H1')
    ws['A1'].value = 'Debt Avalanche Plan \u2014 Highest Interest Rate First'
    ws['A1'].font = TITLE_FONT_FN(theme)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36

    # Row 2: Extra monthly payment (linked to Snowball)
    ws.cell(row=2, column=1, value='Extra Monthly Payment:').font = DATA_FONT_BOLD
    ws.merge_cells('A2:C2')
    cell = ws.cell(row=2, column=4)
    cell.value = "='Snowball Plan'!D2"
    cell.font = INPUT_FONT_BOLD
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    ws.row_dimensions[2].height = 28

    # Row 3: Total monthly payment
    ws.cell(row=3, column=1, value='Total Monthly Payment:').font = DATA_FONT_BOLD
    ws.merge_cells('A3:C3')
    cell = ws.cell(row=3, column=4)
    cell.value = f"={inv}!E{data_start_inv + DEBT_ROWS}+D2"
    cell.font = DATA_FONT_BOLD
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    ws.row_dimensions[3].height = 24

    ws.row_dimensions[4].height = 6

    headers = ["#", "Debt Name", "Balance", "Rate", "Min Payment",
               "Months", "Interest Paid", "Payoff Date"]
    style_header_row(ws, 5, 1, 8, headers, theme)
    ws.freeze_panes = 'A6'

    summary_start = 6
    for i in range(DEBT_ROWS):
        row = summary_start + i
        style_data_row(ws, row, 1, 8, i)
        rank = i + 1
        calc_col = get_column_letter(i + 2)

        ws.cell(row=row, column=1, value=rank).font = DATA_FONT_BOLD
        ws.cell(row=row, column=1).alignment = CENTER

        # B: Debt Name (sorted by LARGE interest rate)
        cell = ws.cell(row=row, column=2)
        cell.value = (
            f'=IFERROR(INDEX({inv}!A{data_start_inv}:A{data_end_inv},'
            f'MATCH(LARGE({inv}!D{data_start_inv}:D{data_end_inv},{rank}),'
            f'{inv}!D{data_start_inv}:D{data_end_inv},0)),"")'
        )
        cell.font = DATA_FONT_BOLD
        cell.alignment = LEFT_INDENT

        # C: Balance
        cell = ws.cell(row=row, column=3)
        cell.value = (
            f'=IFERROR(INDEX({inv}!C{data_start_inv}:C{data_end_inv},'
            f'MATCH(LARGE({inv}!D{data_start_inv}:D{data_end_inv},{rank}),'
            f'{inv}!D{data_start_inv}:D{data_end_inv},0)),"")'
        )
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # D: Interest Rate (sorted descending)
        cell = ws.cell(row=row, column=4)
        cell.value = f'=IFERROR(LARGE({inv}!D{data_start_inv}:D{data_end_inv},{rank}),"")'
        cell.font = DATA_FONT
        cell.number_format = '0.0%'
        cell.alignment = CENTER

        # E: Min Payment
        cell = ws.cell(row=row, column=5)
        cell.value = (
            f'=IFERROR(INDEX({inv}!E{data_start_inv}:E{data_end_inv},'
            f'MATCH(LARGE({inv}!D{data_start_inv}:D{data_end_inv},{rank}),'
            f'{inv}!D{data_start_inv}:D{data_end_inv},0)),"")'
        )
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # F: Months (from calc sheet)
        cell = ws.cell(row=row, column=6)
        cell.value = (
            f'=IF(OR(C{row}="",C{row}=0),"",'
            f'IF({calc}!{calc_col}{last_calc_row}>0,"N/A",'
            f'COUNTIF({calc}!{calc_col}$6:{calc_col}${last_calc_row},">0")))'
        )
        cell.font = DATA_FONT_BOLD
        cell.alignment = CENTER

        # G: Interest Paid
        cell = ws.cell(row=row, column=7)
        cell.value = (
            f'=IF(OR(C{row}="",C{row}=0),"",'
            f'D{row}/12*SUM({calc}!{calc_col}$6:{calc_col}${last_calc_row - 1}))'
        )
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # H: Payoff Date
        cell = ws.cell(row=row, column=8)
        cell.value = (
            f'=IF(OR(F{row}="",F{row}="N/A"),"",'
            f'DATE(YEAR(TODAY()),MONTH(TODAY())+F{row},DAY(TODAY())))'
        )
        cell.font = DATA_FONT
        cell.number_format = 'mmm yyyy'
        cell.alignment = CENTER

    # Totals
    summary_end = summary_start + DEBT_ROWS - 1
    total_row = summary_end + 1
    ws.row_dimensions[total_row].height = 26

    ws.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = accent_fill(theme)
    ws.cell(row=total_row, column=1).alignment = LEFT_INDENT
    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    for col in [2, 4, 5, 6, 8]:
        c = ws.cell(row=total_row, column=col)
        c.fill = accent_fill(theme)
        c.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=3)
    cell.value = f'=SUM(C{summary_start}:C{summary_end})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=7)
    cell.value = f'={calc}!{cum_col_letter}{last_calc_row}'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Comparison: Interest saved vs Snowball
    comp_row = total_row + 2
    ws.row_dimensions[comp_row].height = 28
    ws.merge_cells(f'A{comp_row}:E{comp_row}')
    cell = ws.cell(row=comp_row, column=1, value='Interest Saved vs Snowball Method:')
    cell.font = Font(name='Arial', size=12, bold=True, color=theme["primary"])
    cell.alignment = LEFT

    cell = ws.cell(row=comp_row, column=6)
    ws.merge_cells(f'F{comp_row}:H{comp_row}')
    # Snowball cumul interest - Avalanche cumul interest = savings
    cell.value = (
        f"={snow_calc}!{cum_col_letter}{last_calc_row}"
        f"-{calc}!{cum_col_letter}{last_calc_row}"
    )
    cell.font = Font(name='Arial', size=14, bold=True, color='006100')
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER
    cell.fill = GREEN_FILL

    months_range = f'F{summary_start}:F{summary_end}'
    ws.conditional_formatting.add(months_range,
        CellIsRule(operator='equal', formula=['"N/A"'],
                   fill=RED_FILL, font=Font(color='9C0006', bold=True)))

    setup_print(ws, 8, comp_row, '5:5')
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 4: PAYMENT LOG (unchanged)
# ============================================================

def build_payment_log(wb, theme):
    ws = wb.create_sheet("Payment Log")
    ws.sheet_properties.tabColor = theme["accent"]

    for col, w in enumerate([14, 22, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    headers = ["Date", "Debt Name", "Amount Paid", "Notes"]
    style_header_row(ws, 1, 1, 4, headers, theme)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:D1'

    inv_range = f"'Debt Inventory'!A2:A{1 + DEBT_ROWS}"
    dv_debt = DataValidation(
        type='list', formula1=f'={inv_range}', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid Debt',
        error='Please select a debt from the dropdown'
    )
    ws.add_data_validation(dv_debt)

    for i in range(PAYMENT_LOG_ROWS):
        row = 2 + i
        style_data_row(ws, row, 1, 4, i)
        ws.cell(row=row, column=1).number_format = 'mm/dd/yyyy'
        ws.cell(row=row, column=1).protection = Protection(locked=False)
        ws.cell(row=row, column=1).font = INPUT_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).protection = Protection(locked=False)
        ws.cell(row=row, column=2).font = INPUT_FONT
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=3).alignment = RIGHT
        ws.cell(row=row, column=3).protection = Protection(locked=False)
        ws.cell(row=row, column=3).font = INPUT_FONT
        ws.cell(row=row, column=4).alignment = LEFT
        ws.cell(row=row, column=4).protection = Protection(locked=False)
        ws.cell(row=row, column=4).font = INPUT_FONT

    dv_debt.add(f'B2:B{1 + PAYMENT_LOG_ROWS}')

    total_row = 2 + PAYMENT_LOG_ROWS
    ws.row_dimensions[total_row].height = 26
    ws.cell(row=total_row, column=2, value='TOTAL PAYMENTS').font = TOTAL_FONT
    ws.cell(row=total_row, column=2).fill = accent_fill(theme)
    ws.cell(row=total_row, column=2).alignment = RIGHT
    cell = ws.cell(row=total_row, column=3)
    cell.value = f'=SUM(C2:C{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER
    for col in [1, 2, 4]:
        ws.cell(row=total_row, column=col).border = TOTALS_BORDER
        ws.cell(row=total_row, column=col).fill = accent_fill(theme)

    setup_print(ws, 4, total_row)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 5: INSTRUCTIONS (unchanged)
# ============================================================

def build_instructions(wb, theme):
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = theme["primary"]
    ws.column_dimensions['A'].width = 85

    lines = [
        ("How to Use Your Debt Payoff Tracker", True, 18, theme["primary"]),
        ("", False, 11, None),
        ("Welcome to your FileSmartCo Debt Payoff Tracker!", False, 11, None),
        ("This spreadsheet helps you create a plan to become debt-free.", False, 11, None),
        ("", False, 11, None),
        ("GETTING STARTED", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("1.  Go to the Debt Inventory tab and enter all your debts.", False, 11, None),
        ("     Name, type, current balance, interest rate (APR), and minimum payment.", False, 10, '808080'),
        ("     Blue text = cells you can edit. Black text = auto-calculated formulas.", False, 10, '808080'),
        ("", False, 6, None),
        ("2.  Check the Snowball Plan and Avalanche Plan tabs.", False, 11, None),
        ("     Your debts are automatically sorted and a payoff plan is generated.", False, 10, '808080'),
        ("     Enter your extra monthly payment in the Snowball Plan tab (row 2).", False, 10, '808080'),
        ("", False, 6, None),
        ("3.  Review the Dashboard for your overview.", False, 11, None),
        ("     See total debt, projected payoff date, and method comparison.", False, 10, '808080'),
        ("", False, 6, None),
        ("4.  Log payments in the Payment Log tab.", False, 11, None),
        ("     Track every payment you make toward your debts.", False, 10, '808080'),
        ("", False, 11, None),
        ("SNOWBALL vs AVALANCHE \u2014 Which Method?", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("SNOWBALL METHOD (Smallest Balance First)", True, 11, theme["accent"]),
        ("  \u2022  Pay minimum on all debts, throw extra at the smallest balance", False, 11, None),
        ("  \u2022  When that debt is paid off, roll its payment to the next smallest", False, 11, None),
        ("  \u2022  Pros: Quick wins keep you motivated", False, 11, None),
        ("  \u2022  Cons: May pay more interest overall", False, 11, None),
        ("", False, 6, None),
        ("AVALANCHE METHOD (Highest Interest Rate First)", True, 11, theme["accent"]),
        ("  \u2022  Pay minimum on all debts, throw extra at the highest interest rate", False, 11, None),
        ("  \u2022  When that debt is paid off, roll its payment to the next highest rate", False, 11, None),
        ("  \u2022  Pros: Saves the most money on interest", False, 11, None),
        ("  \u2022  Cons: May take longer to see first debt eliminated", False, 11, None),
        ("", False, 6, None),
        ("Both methods work! Choose the one that fits your personality.", True, 11, '006100'),
        ("", False, 11, None),
        ("TIPS FOR PAYING OFF DEBT FASTER", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("  \u2022  Every extra dollar counts \u2014 even $20/month makes a difference", False, 11, None),
        ("  \u2022  Use windfalls (tax refunds, bonuses) to make lump-sum payments", False, 11, None),
        ("  \u2022  Consider a balance transfer for high-interest credit cards", False, 11, None),
        ("  \u2022  Avoid taking on new debt while paying off existing debt", False, 11, None),
        ("  \u2022  Celebrate milestones \u2014 each debt paid off is a victory!", False, 11, None),
        ("  \u2022  Review your plan monthly and adjust as needed", False, 11, None),
        ("", False, 11, None),
        ("FOR GOOGLE SHEETS USERS", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("1.  Open the link from your download", False, 11, None),
        ("2.  File > Make a copy", False, 11, None),
        ("3.  Save to your Google Drive", False, 11, None),
        ("Do NOT edit the original \u2014 always make a copy first.", True, 11, '9C0006'),
        ("", False, 11, None),
        ("SUPPORT", True, 13, theme["primary"]),
        ("Questions? Message us on Etsy \u2014 we're happy to help!", False, 11, None),
        ("", False, 11, None),
        ("Thank you for choosing FileSmartCo!", True, 13, theme["primary"]),
    ]

    for i, (text, bold, size, color) in enumerate(lines):
        cell = ws.cell(row=i + 1, column=1, value=text)
        cell.font = Font(name='Arial', bold=bold, size=size, color=color or '333333')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 6: DASHBOARD (with fixes)
# ============================================================

def build_dashboard(wb, theme):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = theme["primary"]

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    inv = "'Debt Inventory'"
    snow = "'Snowball Plan'"
    aval = "'Avalanche Plan'"
    plog = "'Payment Log'"
    snow_calc = "'Snowball Calc'"
    aval_calc = "'Avalanche Calc'"
    inv_total_row = 2 + DEBT_ROWS
    snow_total_row = 6 + DEBT_ROWS
    last_calc_row = 6 + MAX_MONTHS
    cum_col_letter = get_column_letter(DEBT_ROWS + 3)

    # Row 1: Title
    ws.merge_cells('A1:H1')
    ws['A1'].value = 'Debt Payoff Dashboard'
    ws['A1'].font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 6

    # KPI Cards
    primary_border = Side(style='thin', color=theme["primary"])
    card_header_border = Border(top=primary_border, left=primary_border,
                                right=primary_border, bottom=primary_border)
    card_spans = [
        ('A3:B3', 1, 'TOTAL DEBT'),
        ('C3:D3', 3, 'MONTHLY PAYMENT'),
        ('E3:F3', 5, 'DEBT-FREE DATE'),
        ('G3:H3', 7, 'ACTIVE DEBTS'),
    ]
    for merge_range, col, label in card_spans:
        ws.merge_cells(merge_range)
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = card_header_border
        c2 = ws.cell(row=3, column=col + 1)
        c2.fill = header_fill(theme)
        c2.border = card_header_border

    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 40

    card_val_border = Border(left=primary_border, right=primary_border,
                             bottom=Side(style='medium', color=theme["accent"]))
    for merge in ['A4:B4', 'C4:D4', 'E4:F4', 'G4:H4']:
        ws.merge_cells(merge)
    for col in range(1, 9):
        c = ws.cell(row=4, column=col)
        c.fill = light_fill(theme)
        c.border = card_val_border

    kpi_font = Font(name='Arial', size=16, bold=True, color='000000')

    # A4: Total Debt
    cell = ws.cell(row=4, column=1)
    cell.value = f"={inv}!C{inv_total_row}"
    cell.font = kpi_font
    cell.number_format = '"$"#,##0'
    cell.alignment = CENTER

    # C4: Monthly Payment
    cell = ws.cell(row=4, column=3)
    cell.value = f"={snow}!D3"
    cell.font = kpi_font
    cell.number_format = '"$"#,##0'
    cell.alignment = CENTER

    # E4: Debt-Free Date — FIX: IF guard for empty state
    cell = ws.cell(row=4, column=5)
    cell.value = (
        f'=IF({inv}!C{inv_total_row}=0,"",'
        f'MAX({snow}!H6:H{5 + DEBT_ROWS}))'
    )
    cell.font = Font(name='Arial', size=14, bold=True, color='000000')
    cell.number_format = 'mmm yyyy'
    cell.alignment = CENTER

    # G4: Active Debts
    cell = ws.cell(row=4, column=7)
    cell.value = f"={inv}!C{inv_total_row + 1}"
    cell.font = kpi_font
    cell.alignment = CENTER

    ws.row_dimensions[5].height = 6

    # Progress section
    ws.merge_cells('A6:H6')
    ws['A6'].value = 'Payoff Progress'
    ws['A6'].font = Font(name='Arial', size=13, bold=True, color=theme["primary"])
    ws['A6'].alignment = LEFT
    ws.row_dimensions[6].height = 26
    for col in range(1, 9):
        ws.cell(row=6, column=col).border = Border(
            bottom=Side(style='medium', color=theme["accent"]))

    ws.merge_cells('A7:B7')
    ws.cell(row=7, column=1, value='Total Payments Made:').font = DATA_FONT_BOLD
    ws.cell(row=7, column=1).alignment = RIGHT
    cell = ws.cell(row=7, column=3)
    cell.value = f"={plog}!C{2 + PAYMENT_LOG_ROWS}"
    cell.font = Font(name='Arial', size=12, bold=True, color='006100')
    cell.number_format = '"$"#,##0.00'
    cell.alignment = LEFT

    ws.merge_cells('A8:B8')
    ws.cell(row=8, column=1, value='Percentage Paid Off:').font = DATA_FONT_BOLD
    ws.cell(row=8, column=1).alignment = RIGHT
    cell = ws.cell(row=8, column=3)
    cell.value = f'=IF(A4=0,"",C7/(C7+A4))'
    cell.font = Font(name='Arial', size=12, bold=True, color=theme["primary"])
    cell.number_format = '0.0%'
    cell.alignment = LEFT

    ws.conditional_formatting.add('C8',
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=1, color=theme["highlight"]))

    ws.row_dimensions[7].height = 28
    ws.row_dimensions[8].height = 28
    ws.row_dimensions[9].height = 6

    # Comparison section
    ws.merge_cells('A10:H10')
    ws['A10'].value = 'Snowball vs Avalanche Comparison'
    ws['A10'].font = Font(name='Arial', size=13, bold=True, color=theme["primary"])
    ws['A10'].alignment = LEFT
    ws.row_dimensions[10].height = 26
    for col in range(1, 9):
        ws.cell(row=10, column=col).border = Border(
            bottom=Side(style='medium', color=theme["accent"]))

    comp_headers = ["", "Snowball", "Avalanche", "Difference"]
    for i, hdr in enumerate(comp_headers):
        col = 2 + i
        cell = ws.cell(row=11, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = header_border(theme)
    ws.row_dimensions[11].height = 28

    # Total Interest comparison (from calc sheets)
    ws.cell(row=12, column=2, value='Total Interest').font = DATA_FONT_BOLD
    ws.cell(row=12, column=2).alignment = RIGHT
    ws.cell(row=12, column=2).border = THIN_BORDER

    cell = ws.cell(row=12, column=3)
    cell.value = f"={snow_calc}!{cum_col_letter}{last_calc_row}"
    cell.font = DATA_FONT
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    cell = ws.cell(row=12, column=4)
    cell.value = f"={aval_calc}!{cum_col_letter}{last_calc_row}"
    cell.font = DATA_FONT
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER
    cell.border = THIN_BORDER

    cell = ws.cell(row=12, column=5)
    cell.value = '=C12-D12'
    cell.font = Font(name='Arial', size=10, bold=True, color='006100')
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    cell.fill = GREEN_FILL

    ws.row_dimensions[12].height = 24
    ws.row_dimensions[13].height = 10

    # Pie Chart
    ws.merge_cells('A14:H14')
    ws['A14'].value = 'Debt Breakdown'
    ws['A14'].font = Font(name='Arial', size=13, bold=True, color=theme["primary"])
    ws['A14'].alignment = LEFT
    ws.row_dimensions[14].height = 26

    pie = PieChart()
    pie.title = 'Debt by Type'
    pie.style = 10
    cats = Reference(wb["Debt Inventory"], min_col=1, min_row=2, max_row=1 + len(SAMPLE_DEBTS))
    vals = Reference(wb["Debt Inventory"], min_col=3, min_row=2, max_row=1 + len(SAMPLE_DEBTS))
    pie.add_data(vals, titles_from_data=False)
    pie.set_categories(cats)
    pie.width = 18
    pie.height = 13

    palette = theme["chart_palette"]
    slices = []
    for i in range(len(SAMPLE_DEBTS)):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = palette[i % len(palette)]
        slices.append(dp)
    pie.series[0].data_points = slices

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showVal = False
    pie.legend.position = 'r'
    pie.legend.overlay = False

    ws.add_chart(pie, 'A15')

    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# MAIN
# ============================================================

def generate_workbook(theme_name, theme):
    wb = Workbook()
    temp = wb.active
    temp.title = '_temp'

    # Build in dependency order:
    # 1. Inventory (data source)
    # 2. Plan tabs (sort formulas reference inventory)
    # 3. Calc sheets (amortization formulas reference plan tabs)
    # 4. Dashboard (references everything)
    build_debt_inventory(wb, theme)
    build_snowball_plan(wb, theme)
    build_avalanche_plan(wb, theme)
    build_method_calc(wb, 'snowball')
    build_method_calc(wb, 'avalanche')
    build_payment_log(wb, theme)
    build_instructions(wb, theme)
    build_dashboard(wb, theme)

    del wb['_temp']

    wb.properties.creator = 'FileSmartCo'
    wb.properties.title = f'Debt Payoff Tracker - {theme_name}'
    wb.properties.company = 'FileSmartCo'

    return wb


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for name, theme in THEMES.items():
        fname = f'Debt-Payoff-Tracker-{name}.xlsx'
        path = os.path.join(OUTPUT_DIR, fname)
        print(f'Generating {fname}...')
        wb = generate_workbook(name, theme)
        wb.save(path)
        print(f'  \u2713 {path}')
    print('\nDone! All 4 variants generated.')


if __name__ == '__main__':
    main()
