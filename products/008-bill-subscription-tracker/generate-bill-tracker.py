"""
#008 Bill & Subscription Tracker Generator
openpyxl. 4 colour themes. 5 tabs. Native charts.
Same pattern as #006 Budget Planner and #007 Debt Payoff Tracker.
"""

import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from colours import THEMES

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- Constants ----------
BILL_ROWS = 30
SUB_ROWS = 25

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

BILL_CATEGORIES = [
    "Housing", "Utilities", "Insurance", "Phone/Internet",
    "Transportation", "Subscriptions", "Debt Payments", "Other",
]
NUM_BILL_CATS = len(BILL_CATEGORIES)

SUB_CATEGORIES = [
    "Streaming", "Software", "Fitness", "Food Delivery",
    "News/Media", "Music", "Gaming", "Cloud Storage", "Other",
]

FREQUENCIES = ["Weekly", "Monthly", "Quarterly", "Semi-Annual", "Annual"]
YES_NO = ["Yes", "No"]
DECISIONS = ["Keep", "Cancel", "Review"]
BILLING_CYCLES = ["Monthly", "Annual"]

# (Name, Category, Amount, DueDay, Frequency, AutoPay, PaidThisMonth)
SAMPLE_BILLS = [
    ("Rent", "Housing", 1800, 1, "Monthly", "Yes", "Yes"),
    ("Electricity", "Utilities", 120, 15, "Monthly", "No", "No"),
    ("Internet", "Phone/Internet", 79.99, 5, "Monthly", "Yes", "No"),
    ("Car Insurance", "Insurance", 145, 20, "Quarterly", "Yes", "No"),
    ("Phone", "Phone/Internet", 65, 12, "Monthly", "Yes", "Yes"),
    ("Water", "Utilities", 45, 25, "Monthly", "No", "No"),
]

# (Name, Category, MonthlyCost, BillingCycle, RenewalDate, CancelLink, Decision)
SAMPLE_SUBS = [
    ("Netflix", "Streaming", 15.99, "Monthly", date(2026, 4, 15), "", "Keep"),
    ("Spotify", "Music", 12.99, "Monthly", date(2026, 4, 1), "", "Keep"),
    ("ChatGPT Plus", "Software", 20.00, "Monthly", date(2026, 3, 20), "", "Keep"),
    ("Adobe Creative Cloud", "Software", 54.99, "Monthly", date(2026, 5, 1), "", "Review"),
    ("Gym Membership", "Fitness", 49.99, "Monthly", date(2026, 6, 1), "", "Keep"),
    ("Amazon Prime", "Streaming", 14.99, "Annual", date(2026, 9, 15), "", "Keep"),
    ("iCloud Storage", "Cloud Storage", 2.99, "Monthly", date(2026, 3, 10), "", "Keep"),
    ("DoorDash DashPass", "Food Delivery", 9.99, "Monthly", date(2026, 4, 5), "", "Review"),
]

OUTPUT_DIR = "/mnt/c/Users/jayso/businesses/etsy-templates/products/v2/output/008-bill-subscription-tracker"


# ---------- Reusable styles (from #006/#007 pattern) ----------

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10, color='333333')
DATA_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='333333')
TOTAL_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')
INPUT_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='0000FF')

ZEBRA_FILL = PatternFill(fill_type='solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF')
GREEN_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')
RED_FILL = PatternFill(fill_type='solid', fgColor='FFC7CE')
YELLOW_FILL = PatternFill(fill_type='solid', fgColor='FFEB9C')


def header_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["primary"])

def accent_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["accent"])

def light_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["light_bg"])


THIN_SIDE = Side(style='thin', color='D9D9D9')
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, bottom=THIN_SIDE, right=THIN_SIDE)
TOTALS_BORDER = Border(top=Side(style='double'), bottom=Side(style='double'),
                        left=THIN_SIDE, right=THIN_SIDE)

def header_border(theme):
    return Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE,
                  bottom=Side(style='medium', color=theme["primary"]))

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
LEFT_INDENT = Alignment(horizontal='left', vertical='center', indent=1)
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ---------- Helpers ----------

def style_header_row(ws, row, start_col, end_col, headers, theme):
    for i, hdr in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = header_border(theme)
    ws.row_dimensions[row].height = 28


def style_data_row(ws, row, start_col, end_col, row_idx):
    ws.row_dimensions[row].height = 22
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.fill = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL


def setup_print(ws, num_cols, last_row, header_rows='1:1'):
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_area = f'A1:{get_column_letter(num_cols)}{last_row}'
    ws.print_title_rows = header_rows


# ============================================================
# TAB 2: BILL TRACKER
# ============================================================

def build_bill_tracker(wb, theme):
    ws = wb.create_sheet("Bill Tracker")
    ws.sheet_properties.tabColor = theme["accent"]

    # Column widths: A=Bill Name, B=Category, C=Amount, D=Due Day,
    #   E=Frequency, F=Auto-Pay, G=Paid, H=Status, I=Monthly Cost
    for col, w in enumerate([24, 18, 14, 10, 14, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    headers = ["Bill Name", "Category", "Amount", "Due Day",
               "Frequency", "Auto-Pay", "Paid This Month", "Status", "Monthly Cost"]
    style_header_row(ws, 1, 1, 9, headers, theme)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:I1'

    # Data validations
    dv_cat = DataValidation(
        type='list', formula1='"' + ','.join(BILL_CATEGORIES) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Category', error='Please select from the dropdown list')
    dv_freq = DataValidation(
        type='list', formula1='"' + ','.join(FREQUENCIES) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Frequency', error='Please select from the dropdown list')
    dv_yn_auto = DataValidation(
        type='list', formula1='"' + ','.join(YES_NO) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid', error='Please select Yes or No')
    dv_yn_paid = DataValidation(
        type='list', formula1='"' + ','.join(YES_NO) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid', error='Please select Yes or No')
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_freq)
    ws.add_data_validation(dv_yn_auto)
    ws.add_data_validation(dv_yn_paid)

    data_start = 2
    for i in range(BILL_ROWS):
        row = data_start + i
        style_data_row(ws, row, 1, 9, i)

        # A: Bill Name (unlocked, blue input)
        ws.cell(row=row, column=1).font = INPUT_FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_INDENT
        ws.cell(row=row, column=1).protection = Protection(locked=False)

        # B: Category (unlocked, blue input)
        ws.cell(row=row, column=2).font = INPUT_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).protection = Protection(locked=False)

        # C: Amount (unlocked, blue input)
        cell = ws.cell(row=row, column=3)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # D: Due Day (unlocked, blue input, 1-31)
        cell = ws.cell(row=row, column=4)
        cell.font = INPUT_FONT
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)

        # E: Frequency (unlocked, dropdown)
        ws.cell(row=row, column=5).font = INPUT_FONT
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=5).protection = Protection(locked=False)

        # F: Auto-Pay (unlocked, dropdown)
        ws.cell(row=row, column=6).font = INPUT_FONT
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=6).protection = Protection(locked=False)

        # G: Paid This Month (unlocked, dropdown)
        ws.cell(row=row, column=7).font = INPUT_FONT
        ws.cell(row=row, column=7).alignment = CENTER
        ws.cell(row=row, column=7).protection = Protection(locked=False)

        # H: Status (formula — locked)
        cell = ws.cell(row=row, column=8)
        cell.value = (
            f'=IF(A{row}="","",IF(G{row}="Yes","Paid",'
            f'IF(AND(D{row}<>"",D{row}<DAY(TODAY())),"Overdue",'
            f'IF(AND(D{row}<>"",D{row}-DAY(TODAY())<=7),"Due Soon","Upcoming"))))'
        )
        cell.font = DATA_FONT_BOLD
        cell.alignment = CENTER

        # I: Monthly Cost (formula — locked)
        cell = ws.cell(row=row, column=9)
        cell.value = (
            f'=IF(A{row}="","",IF(E{row}="Weekly",C{row}*4.33,'
            f'IF(E{row}="Quarterly",C{row}/3,'
            f'IF(E{row}="Semi-Annual",C{row}/6,'
            f'IF(E{row}="Annual",C{row}/12,C{row})))))'
        )
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # Fill sample data
        if i < len(SAMPLE_BILLS):
            name, cat, amt, day, freq, auto, paid = SAMPLE_BILLS[i]
            ws.cell(row=row, column=1).value = name
            ws.cell(row=row, column=2).value = cat
            ws.cell(row=row, column=3).value = amt
            ws.cell(row=row, column=4).value = day
            ws.cell(row=row, column=5).value = freq
            ws.cell(row=row, column=6).value = auto
            ws.cell(row=row, column=7).value = paid

    dv_cat.add(f'B{data_start}:B{data_start + BILL_ROWS - 1}')
    dv_freq.add(f'E{data_start}:E{data_start + BILL_ROWS - 1}')
    dv_yn_auto.add(f'F{data_start}:F{data_start + BILL_ROWS - 1}')
    dv_yn_paid.add(f'G{data_start}:G{data_start + BILL_ROWS - 1}')

    # Due Day validation (1-31)
    dv_day = DataValidation(
        type='whole', operator='between', formula1='1', formula2='31',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Day', error='Please enter a day between 1 and 31')
    ws.add_data_validation(dv_day)
    dv_day.add(f'D{data_start}:D{data_start + BILL_ROWS - 1}')

    # Totals row
    total_row = data_start + BILL_ROWS
    ws.row_dimensions[total_row].height = 26

    ws.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = accent_fill(theme)
    ws.cell(row=total_row, column=1).alignment = LEFT_INDENT
    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    for col in [2, 4, 5, 6, 7, 8]:
        c = ws.cell(row=total_row, column=col)
        c.fill = accent_fill(theme)
        c.border = TOTALS_BORDER

    # Total Amount
    cell = ws.cell(row=total_row, column=3)
    cell.value = f'=SUM(C{data_start}:C{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Total Monthly Cost
    cell = ws.cell(row=total_row, column=9)
    cell.value = f'=SUM(I{data_start}:I{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Conditional formatting: Status column
    status_range = f'H{data_start}:H{total_row - 1}'
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"Paid"'],
                   fill=GREEN_FILL, font=Font(color='006100', bold=True)))
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"Due Soon"'],
                   fill=YELLOW_FILL, font=Font(color='9C5700', bold=True)))
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"Overdue"'],
                   fill=RED_FILL, font=Font(color='9C0006', bold=True)))
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator='equal', formula=['"Upcoming"'],
                   fill=PatternFill(fill_type='solid', fgColor='E8F0FE'),
                   font=Font(color='1B365D', bold=True)))

    setup_print(ws, 9, total_row)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 3: SUBSCRIPTION TRACKER
# ============================================================

def build_subscription_tracker(wb, theme):
    ws = wb.create_sheet("Subscription Tracker")
    ws.sheet_properties.tabColor = theme["highlight"]

    # A=Service, B=Category, C=Monthly, D=Annual, E=Cycle, F=Renewal, G=Cancel Link, H=Decision
    for col, w in enumerate([24, 18, 14, 14, 14, 14, 28, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    headers = ["Service Name", "Category", "Monthly Cost", "Annual Cost",
               "Billing Cycle", "Renewal Date", "Cancellation Link", "Decision"]
    style_header_row(ws, 1, 1, 8, headers, theme)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:H1'

    # Data validations
    dv_cat = DataValidation(
        type='list', formula1='"' + ','.join(SUB_CATEGORIES) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Category', error='Select from dropdown')
    dv_cycle = DataValidation(
        type='list', formula1='"' + ','.join(BILLING_CYCLES) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Cycle', error='Select Monthly or Annual')
    dv_decision = DataValidation(
        type='list', formula1='"' + ','.join(DECISIONS) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Decision', error='Select Keep, Cancel, or Review')
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_cycle)
    ws.add_data_validation(dv_decision)

    data_start = 2
    for i in range(SUB_ROWS):
        row = data_start + i
        style_data_row(ws, row, 1, 8, i)

        # A: Service Name (unlocked)
        ws.cell(row=row, column=1).font = INPUT_FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_INDENT
        ws.cell(row=row, column=1).protection = Protection(locked=False)

        # B: Category (unlocked)
        ws.cell(row=row, column=2).font = INPUT_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).protection = Protection(locked=False)

        # C: Monthly Cost (unlocked)
        cell = ws.cell(row=row, column=3)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # D: Annual Cost (formula = Monthly × 12)
        cell = ws.cell(row=row, column=4)
        cell.value = f'=IF(C{row}="","",C{row}*12)'
        cell.font = DATA_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # E: Billing Cycle (unlocked)
        ws.cell(row=row, column=5).font = INPUT_FONT
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=5).protection = Protection(locked=False)

        # F: Renewal Date (unlocked)
        cell = ws.cell(row=row, column=6)
        cell.font = INPUT_FONT
        cell.number_format = 'mm/dd/yyyy'
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)

        # G: Cancellation Link (unlocked)
        ws.cell(row=row, column=7).font = INPUT_FONT
        ws.cell(row=row, column=7).alignment = LEFT
        ws.cell(row=row, column=7).protection = Protection(locked=False)

        # H: Decision (unlocked)
        ws.cell(row=row, column=8).font = INPUT_FONT
        ws.cell(row=row, column=8).alignment = CENTER
        ws.cell(row=row, column=8).protection = Protection(locked=False)

        # Fill sample data
        if i < len(SAMPLE_SUBS):
            name, cat, monthly, cycle, renewal, link, decision = SAMPLE_SUBS[i]
            ws.cell(row=row, column=1).value = name
            ws.cell(row=row, column=2).value = cat
            ws.cell(row=row, column=3).value = monthly
            ws.cell(row=row, column=5).value = cycle
            ws.cell(row=row, column=6).value = renewal
            ws.cell(row=row, column=7).value = link
            ws.cell(row=row, column=8).value = decision

    dv_cat.add(f'B{data_start}:B{data_start + SUB_ROWS - 1}')
    dv_cycle.add(f'E{data_start}:E{data_start + SUB_ROWS - 1}')
    dv_decision.add(f'H{data_start}:H{data_start + SUB_ROWS - 1}')

    # Totals row
    total_row = data_start + SUB_ROWS
    ws.row_dimensions[total_row].height = 26

    ws.cell(row=total_row, column=1, value='TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = accent_fill(theme)
    ws.cell(row=total_row, column=1).alignment = LEFT_INDENT
    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    for col in [2, 5, 6, 7, 8]:
        c = ws.cell(row=total_row, column=col)
        c.fill = accent_fill(theme)
        c.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=3)
    cell.value = f'=SUM(C{data_start}:C{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=4)
    cell.value = f'=SUM(D{data_start}:D{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = accent_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Decision summary section
    summary_row = total_row + 2
    ws.row_dimensions[summary_row].height = 24
    ws.cell(row=summary_row, column=1, value='Decision Summary').font = Font(
        name='Arial', size=12, bold=True, color=theme["primary"])
    ws.cell(row=summary_row, column=1).alignment = LEFT

    for j, (label, decision) in enumerate([("Keep", "Keep"), ("Cancel", "Cancel"), ("Review", "Review")]):
        col_label = 1 + j * 3  # A, D, G
        col_count = 2 + j * 3  # B, E, H
        r = summary_row + 1
        ws.row_dimensions[r].height = 22

        ws.cell(row=r, column=col_label, value=f'{label}:').font = DATA_FONT_BOLD
        ws.cell(row=r, column=col_label).alignment = RIGHT

        cell = ws.cell(row=r, column=col_count)
        cell.value = f'=COUNTIF(H{data_start}:H{total_row - 1},"{decision}")'
        cell.font = Font(name='Arial', size=12, bold=True, color=theme["primary"])
        cell.alignment = LEFT

    # Conditional formatting: Decision column
    decision_range = f'H{data_start}:H{total_row - 1}'
    ws.conditional_formatting.add(
        decision_range,
        CellIsRule(operator='equal', formula=['"Keep"'],
                   fill=GREEN_FILL, font=Font(color='006100', bold=True)))
    ws.conditional_formatting.add(
        decision_range,
        CellIsRule(operator='equal', formula=['"Cancel"'],
                   fill=RED_FILL, font=Font(color='9C0006', bold=True, strikethrough=True)))
    ws.conditional_formatting.add(
        decision_range,
        CellIsRule(operator='equal', formula=['"Review"'],
                   fill=YELLOW_FILL, font=Font(color='9C5700', bold=True)))

    # Row-level conditional formatting (whole row tint based on decision)
    data_range = f'A{data_start}:G{total_row - 1}'
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'$H{data_start}="Cancel"'],
                    fill=RED_FILL, font=Font(color='9C0006', strikethrough=True)))
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=[f'$H{data_start}="Review"'],
                    fill=YELLOW_FILL))

    setup_print(ws, 8, summary_row + 1)
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 4: MONTHLY CALENDAR
# ============================================================

def build_monthly_calendar(wb, theme):
    ws = wb.create_sheet("Monthly Calendar")
    ws.sheet_properties.tabColor = theme["accent"]

    # A=Bill Name, B-M=Jan-Dec, N=Annual Total
    ws.column_dimensions['A'].width = 22
    for col in range(2, 14):  # B-M
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions['N'].width = 14

    # Row 1: Title
    ws.merge_cells('A1:N1')
    ws['A1'].value = 'Monthly Bill Calendar \u2014 At a Glance'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color=theme["primary"])
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.row_dimensions[2].height = 4  # spacer

    # Row 3: Headers
    cal_headers = ["Bill Name"] + MONTHS + ["Annual Total"]
    style_header_row(ws, 3, 1, 14, cal_headers, theme)

    ws.freeze_panes = 'B4'

    bt = "'Bill Tracker'"
    data_start = 4  # Calendar data starts at row 4
    bt_data_start = 2  # Bill Tracker data starts at row 2

    # Zero-hiding number format
    zero_hide_fmt = '"$"#,##0;-"$"#,##0;""'

    for i in range(BILL_ROWS):
        cal_row = data_start + i
        bt_row = bt_data_start + i
        style_data_row(ws, cal_row, 1, 14, i)

        # A: Bill Name (linked from Bill Tracker)
        cell = ws.cell(row=cal_row, column=1)
        cell.value = f'=IF({bt}!A{bt_row}="","",' f'{bt}!A{bt_row})'
        cell.font = DATA_FONT_BOLD
        cell.alignment = LEFT_INDENT

        # B-M: Monthly amounts based on frequency
        for m in range(1, 13):
            col = m + 1  # B=2 for Jan(1), C=3 for Feb(2), ...
            cell = ws.cell(row=cal_row, column=col)
            cell.value = (
                f'=IF({bt}!A{bt_row}="",0,'
                f'IF(OR({bt}!E{bt_row}="Monthly",{bt}!E{bt_row}="Weekly"),'
                f'IF({bt}!E{bt_row}="Weekly",{bt}!C{bt_row}*4,{bt}!C{bt_row}),'
                f'IF(AND({bt}!E{bt_row}="Quarterly",OR({m}=1,{m}=4,{m}=7,{m}=10)),{bt}!C{bt_row},'
                f'IF(AND({bt}!E{bt_row}="Semi-Annual",OR({m}=1,{m}=7)),{bt}!C{bt_row},'
                f'IF(AND({bt}!E{bt_row}="Annual",{m}=1),{bt}!C{bt_row},0)))))'
            )
            cell.number_format = zero_hide_fmt
            cell.alignment = RIGHT
            cell.font = DATA_FONT

        # N: Annual Total
        cell = ws.cell(row=cal_row, column=14)
        cell.value = f'=SUM(B{cal_row}:M{cal_row})'
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.font = DATA_FONT_BOLD

    # Total row
    total_row = data_start + BILL_ROWS
    ws.row_dimensions[total_row].height = 26

    ws.cell(row=total_row, column=1, value='MONTHLY TOTAL').font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = accent_fill(theme)
    ws.cell(row=total_row, column=1).alignment = LEFT_INDENT
    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    for col in range(2, 15):
        letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM({letter}{data_start}:{letter}{total_row - 1})'
        cell.font = TOTAL_FONT
        cell.fill = accent_fill(theme)
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.border = TOTALS_BORDER

    # Note row
    note_row = total_row + 2
    ws.merge_cells(f'A{note_row}:N{note_row}')
    cell = ws.cell(row=note_row, column=1)
    cell.value = (
        'Note: Quarterly bills shown in Jan/Apr/Jul/Oct. '
        'Semi-Annual in Jan/Jul. Annual in Jan. '
        'Weekly bills shown as approximate monthly (×4).'
    )
    cell.font = Font(name='Arial', size=9, italic=True, color='808080')
    cell.alignment = WRAP

    setup_print(ws, 14, total_row, '3:3')
    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 5: INSTRUCTIONS
# ============================================================

def build_instructions(wb, theme):
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = theme["primary"]
    ws.column_dimensions['A'].width = 85

    lines = [
        ("How to Use Your Bill & Subscription Tracker", True, 18, theme["primary"]),
        ("", False, 11, None),
        ("Welcome to your FileSmartCo Bill & Subscription Tracker!", False, 11, None),
        ("This spreadsheet helps you manage recurring bills, audit subscriptions,", False, 11, None),
        ("and see exactly where your money goes each month.", False, 11, None),
        ("", False, 11, None),
        ("GETTING STARTED", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("1.  Go to the Bill Tracker tab and enter your recurring bills.", False, 11, None),
        ("     Name, category, amount, due day (1-31), frequency, auto-pay status.", False, 10, '808080'),
        ("     Blue text = cells you can edit. Black text = auto-calculated.", False, 10, '808080'),
        ("", False, 6, None),
        ("2.  Go to the Subscription Tracker and list all your subscriptions.", False, 11, None),
        ("     Include monthly cost, billing cycle, renewal date, and cancellation link.", False, 10, '808080'),
        ("     Set a Decision for each: Keep, Cancel, or Review.", False, 10, '808080'),
        ("", False, 6, None),
        ("3.  Check the Monthly Calendar for a year-at-a-glance view.", False, 11, None),
        ("     See which bills hit each month and your total monthly outflow.", False, 10, '808080'),
        ("", False, 6, None),
        ("4.  Review the Dashboard for your complete overview.", False, 11, None),
        ("     KPI cards, spending by category chart, and monthly trends.", False, 10, '808080'),
        ("", False, 11, None),
        ("HOW TO ENTER BILLS", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("\u2022  Amount = what you pay each billing period (not monthly equivalent)", False, 11, None),
        ("\u2022  Due Day = day of month (1-31) when payment is due", False, 11, None),
        ("\u2022  Monthly Cost column auto-calculates the monthly equivalent", False, 11, None),
        ("\u2022  Update 'Paid This Month' at the start of each month (reset all to No)", False, 11, None),
        ("\u2022  Mark bills as paid throughout the month \u2014 Status updates automatically", False, 11, None),
        ("", False, 11, None),
        ("THE SUBSCRIPTION AUDIT \u2014 The 3-Question Test", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("For each subscription, ask yourself:", False, 11, None),
        ("", False, 6, None),
        ("  1.  Did I actually use this in the last 30 days?", True, 11, theme["accent"]),
        ("       If not, it's a strong cancel candidate.", False, 10, '808080'),
        ("", False, 6, None),
        ("  2.  Would I buy this again at this price right now?", True, 11, theme["accent"]),
        ("       If you hesitate, mark it for Review.", False, 10, '808080'),
        ("", False, 6, None),
        ("  3.  Can I get this cheaper or free elsewhere?", True, 11, theme["accent"]),
        ("       Check for free tiers, student/family plans, or bundle deals.", False, 10, '808080'),
        ("", False, 6, None),
        ("Mark each subscription as Keep, Cancel, or Review in the Decision column.", False, 11, None),
        ("Cancelled subscriptions will show with red strikethrough.", False, 11, None),
        ("", False, 11, None),
        ("TIPS FOR NEGOTIATING BILLS DOWN", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("\u2022  Call your provider and ask for their best current rate", False, 11, None),
        ("\u2022  Mention competitor offers \u2014 retention departments have authority to discount", False, 11, None),
        ("\u2022  Bundle services when possible (internet + phone, insurance packages)", False, 11, None),
        ("\u2022  Ask about autopay discounts (many providers offer 5-10% off)", False, 11, None),
        ("\u2022  Review insurance annually \u2014 rates change and loyalty doesn't always pay", False, 11, None),
        ("\u2022  Negotiate medical bills \u2014 providers often accept 60-80% of the billed amount", False, 11, None),
        ("", False, 11, None),
        ("MONTHLY ROUTINE", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("  Start of month:  Reset all 'Paid This Month' to No", False, 11, None),
        ("  Throughout:      Mark bills as Paid when you pay them", False, 11, None),
        ("  End of month:    Review Dashboard, audit subscriptions quarterly", False, 11, None),
        ("", False, 11, None),
        ("FOR GOOGLE SHEETS USERS", True, 13, theme["primary"]),
        ("", False, 6, None),
        ("1.  Open the link from your download", False, 11, None),
        ("2.  File > Make a copy", False, 11, None),
        ("3.  Save to your Google Drive", False, 11, None),
        ("Do NOT edit the original \u2014 always make a copy first.", True, 11, '9C0006'),
        ("", False, 11, None),
        ("SUPPORT", True, 13, theme["primary"]),
        ("Questions? Message us on Etsy \u2014 we're happy to help!", False, 11, None),
        ("", False, 11, None),
        ("Thank you for choosing FileSmartCo!", True, 13, theme["primary"]),
    ]

    for i, (text, bold, size, color) in enumerate(lines):
        cell = ws.cell(row=i + 1, column=1, value=text)
        cell.font = Font(name='Arial', bold=bold, size=size, color=color or '333333')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# TAB 1: DASHBOARD (built last — references all other tabs)
# ============================================================

def build_dashboard(wb, theme):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = theme["primary"]

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    # Helper data columns (to the right)
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 14

    bt = "'Bill Tracker'"
    st = "'Subscription Tracker'"
    cal = "'Monthly Calendar'"
    bt_total_row = 2 + BILL_ROWS
    st_total_row = 2 + SUB_ROWS
    cal_total_row = 4 + BILL_ROWS

    # Row 1: Title
    ws.merge_cells('A1:H1')
    ws['A1'].value = 'Bill & Subscription Dashboard'
    ws['A1'].font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 6

    # ---- KPI Cards ----
    primary_border_side = Side(style='thin', color=theme["primary"])
    card_header_border = Border(top=primary_border_side, left=primary_border_side,
                                right=primary_border_side, bottom=primary_border_side)

    card_spans = [
        ('A3:B3', 1, 'TOTAL MONTHLY BILLS'),
        ('C3:D3', 3, 'TOTAL ANNUAL COST'),
        ('E3:F3', 5, 'ACTIVE SUBSCRIPTIONS'),
        ('G3:H3', 7, 'DUE THIS WEEK'),
    ]
    for merge_range, col, label in card_spans:
        ws.merge_cells(merge_range)
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = card_header_border
        c2 = ws.cell(row=3, column=col + 1)
        c2.fill = header_fill(theme)
        c2.border = card_header_border

    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 40

    card_val_border = Border(
        left=primary_border_side, right=primary_border_side,
        bottom=Side(style='medium', color=theme["accent"]))

    for merge in ['A4:B4', 'C4:D4', 'E4:F4', 'G4:H4']:
        ws.merge_cells(merge)
    for col in range(1, 9):
        c = ws.cell(row=4, column=col)
        c.fill = light_fill(theme)
        c.border = card_val_border

    kpi_font = Font(name='Arial', size=16, bold=True, color='000000')

    # A4: Total Monthly Bills (from Bill Tracker Monthly Cost total)
    cell = ws.cell(row=4, column=1)
    cell.value = f"={bt}!I{bt_total_row}"
    cell.font = kpi_font
    cell.number_format = '"$"#,##0'
    cell.alignment = CENTER

    # C4: Total Annual Cost = Monthly Bills × 12 + Annual Subscriptions
    cell = ws.cell(row=4, column=3)
    cell.value = f"=A4*12+{st}!D{st_total_row}"
    cell.font = kpi_font
    cell.number_format = '"$"#,##0'
    cell.alignment = CENTER

    # E4: Active Subscriptions (count non-empty names)
    cell = ws.cell(row=4, column=5)
    cell.value = f'=COUNTA({st}!A2:A{st_total_row - 1})'
    cell.font = kpi_font
    cell.alignment = CENTER

    # G4: Upcoming Due This Week (unpaid bills due within 7 days)
    cell = ws.cell(row=4, column=7)
    cell.value = (
        f'=SUMPRODUCT(({bt}!A$2:A${bt_total_row - 1}<>"")'
        f'*({bt}!G$2:G${bt_total_row - 1}<>"Yes")'
        f'*({bt}!D$2:D${bt_total_row - 1}>=DAY(TODAY()))'
        f'*({bt}!D$2:D${bt_total_row - 1}<=DAY(TODAY())+7))'
    )
    cell.font = kpi_font
    cell.alignment = CENTER

    ws.row_dimensions[5].height = 6

    # ---- Section: Spending by Category ----
    ws.merge_cells('A6:H6')
    ws['A6'].value = 'Spending by Category'
    ws['A6'].font = Font(name='Arial', size=13, bold=True, color=theme["primary"])
    ws['A6'].alignment = LEFT
    ws.row_dimensions[6].height = 26
    for col in range(1, 9):
        ws.cell(row=6, column=col).border = Border(
            bottom=Side(style='medium', color=theme["accent"]))

    # ---- Helper data: Pie chart (Category breakdown) ----
    ws.cell(row=3, column=10, value='Category').font = Font(name='Arial', size=9, bold=True, color='808080')
    ws.cell(row=3, column=11, value='Monthly $').font = Font(name='Arial', size=9, bold=True, color='808080')

    for idx, cat in enumerate(BILL_CATEGORIES):
        r = 4 + idx
        ws.cell(row=r, column=10, value=cat).font = Font(name='Arial', size=9, color='808080')
        cell = ws.cell(row=r, column=11)
        cell.value = f'=SUMIF({bt}!B$2:B${bt_total_row - 1},J{r},{bt}!I$2:I${bt_total_row - 1})'
        cell.font = Font(name='Arial', size=9, color='808080')
        cell.number_format = '"$"#,##0.00'

    # Pie Chart
    pie = PieChart()
    pie.title = 'Monthly Bills by Category'
    pie.style = 10

    cats = Reference(ws, min_col=10, min_row=4, max_row=3 + NUM_BILL_CATS)
    vals = Reference(ws, min_col=11, min_row=4, max_row=3 + NUM_BILL_CATS)
    pie.add_data(vals, titles_from_data=False)
    pie.set_categories(cats)
    pie.width = 18
    pie.height = 13

    palette = theme["chart_palette"]
    slices = []
    for i in range(NUM_BILL_CATS):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = palette[i % len(palette)]
        slices.append(dp)
    pie.series[0].data_points = slices

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = False
    pie.dataLabels.showCatName = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showVal = False
    pie.legend.position = 'r'
    pie.legend.overlay = False

    ws.add_chart(pie, 'A7')

    # ---- Section: Monthly Spending Trend ----
    ws.merge_cells('A22:H22')
    ws['A22'].value = 'Monthly Bill Spending Trend'
    ws['A22'].font = Font(name='Arial', size=13, bold=True, color=theme["primary"])
    ws['A22'].alignment = LEFT
    ws.row_dimensions[22].height = 26
    for col in range(1, 9):
        ws.cell(row=22, column=col).border = Border(
            bottom=Side(style='medium', color=theme["accent"]))

    # Helper data: Bar chart (monthly totals from Calendar)
    ws.cell(row=15, column=10, value='Month').font = Font(name='Arial', size=9, bold=True, color='808080')
    ws.cell(row=15, column=11, value='Total').font = Font(name='Arial', size=9, bold=True, color='808080')

    cal_ws = wb["Monthly Calendar"]
    for idx, month_name in enumerate(MONTHS):
        r = 16 + idx
        cal_col_letter = get_column_letter(idx + 2)  # B=Jan, C=Feb, ...
        ws.cell(row=r, column=10, value=month_name).font = Font(name='Arial', size=9, color='808080')
        cell = ws.cell(row=r, column=11)
        cell.value = f"={cal}!{cal_col_letter}{cal_total_row}"
        cell.font = Font(name='Arial', size=9, color='808080')
        cell.number_format = '"$"#,##0'

    # Bar Chart
    bar = BarChart()
    bar.type = 'col'
    bar.grouping = 'clustered'
    bar.title = 'Expected Monthly Bill Payments'
    bar.style = 10
    bar.y_axis.title = 'Amount ($)'
    bar.y_axis.numFmt = '"$"#,##0'

    bar_data = Reference(ws, min_col=11, min_row=15, max_row=27)
    bar_cats = Reference(ws, min_col=10, min_row=16, max_row=27)

    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.x_axis.tickLblPos = 'low'
    bar.x_axis.delete = False

    bar.series[0].graphicalProperties.solidFill = theme["chart_budget"]

    bar.width = 22
    bar.height = 13
    bar.legend = None  # single series, no legend needed

    ws.add_chart(bar, 'A23')

    # ---- Subscription Summary Section ----
    ws.merge_cells('A38:H38')
    ws['A38'].value = 'Subscription Overview'
    ws['A38'].font = Font(name='Arial', size=13, bold=True, color=theme["primary"])
    ws['A38'].alignment = LEFT
    ws.row_dimensions[38].height = 26
    for col in range(1, 9):
        ws.cell(row=38, column=col).border = Border(
            bottom=Side(style='medium', color=theme["accent"]))

    # Sub summary cards
    sub_labels = [
        ('A39:B39', 1, 'MONTHLY SUBS'),
        ('C39:D39', 3, 'ANNUAL SUBS'),
        ('E39:F39', 5, 'TO CANCEL'),
        ('G39:H39', 7, 'TO REVIEW'),
    ]
    for merge_range, col, label in sub_labels:
        ws.merge_cells(merge_range)
        cell = ws.cell(row=39, column=col, value=label)
        cell.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = card_header_border
        c2 = ws.cell(row=39, column=col + 1)
        c2.fill = header_fill(theme)
        c2.border = card_header_border
    ws.row_dimensions[39].height = 26

    for merge in ['A40:B40', 'C40:D40', 'E40:F40', 'G40:H40']:
        ws.merge_cells(merge)
    for col in range(1, 9):
        c = ws.cell(row=40, column=col)
        c.fill = light_fill(theme)
        c.border = card_val_border
    ws.row_dimensions[40].height = 36

    sub_font = Font(name='Arial', size=14, bold=True, color='000000')

    # A40: Monthly Sub Total
    cell = ws.cell(row=40, column=1)
    cell.value = f"={st}!C{st_total_row}"
    cell.font = sub_font
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # C40: Annual Sub Total
    cell = ws.cell(row=40, column=3)
    cell.value = f"={st}!D{st_total_row}"
    cell.font = sub_font
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # E40: Cancel Count
    cell = ws.cell(row=40, column=5)
    cell.value = f'=COUNTIF({st}!H$2:H${st_total_row - 1},"Cancel")'
    cell.font = Font(name='Arial', size=14, bold=True, color='9C0006')
    cell.alignment = CENTER

    # G40: Review Count
    cell = ws.cell(row=40, column=7)
    cell.value = f'=COUNTIF({st}!H$2:H${st_total_row - 1},"Review")'
    cell.font = Font(name='Arial', size=14, bold=True, color='9C5700')
    cell.alignment = CENTER

    # Potential savings row
    savings_row = 42
    ws.merge_cells(f'A{savings_row}:D{savings_row}')
    ws.cell(row=savings_row, column=1, value='Potential Monthly Savings (Cancel + Review):').font = (
        Font(name='Arial', size=11, bold=True, color=theme["primary"]))
    ws.cell(row=savings_row, column=1).alignment = RIGHT

    cell = ws.cell(row=savings_row, column=5)
    ws.merge_cells(f'E{savings_row}:F{savings_row}')
    cell.value = (
        f'=SUMPRODUCT(({st}!A$2:A${st_total_row - 1}<>"")'
        f'*(OR({st}!H$2:H${st_total_row - 1}="Cancel",{st}!H$2:H${st_total_row - 1}="Review"))'  # won't work with OR in SUMPRODUCT
        f'*{st}!C$2:C${st_total_row - 1})'
    )
    # Fix: OR doesn't work in SUMPRODUCT arrays. Use addition instead.
    cell.value = (
        f'=SUMPRODUCT(({st}!A$2:A${st_total_row - 1}<>"")'
        f'*(({st}!H$2:H${st_total_row - 1}="Cancel")+({st}!H$2:H${st_total_row - 1}="Review"))'
        f'*{st}!C$2:C${st_total_row - 1})'
    )
    cell.font = Font(name='Arial', size=14, bold=True, color='006100')
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER
    cell.fill = GREEN_FILL
    ws.row_dimensions[savings_row].height = 30

    ws.protection.sheet = True
    ws.protection.enable()


# ============================================================
# MAIN
# ============================================================

def generate_workbook(theme_name, theme):
    wb = Workbook()
    temp = wb.active
    temp.title = '_temp'

    # Build in dependency order:
    # 1. Bill Tracker (data source)
    # 2. Subscription Tracker (data source)
    # 3. Monthly Calendar (references Bill Tracker)
    # 4. Instructions (standalone)
    # 5. Dashboard (references everything, built last)
    build_bill_tracker(wb, theme)
    build_subscription_tracker(wb, theme)
    build_monthly_calendar(wb, theme)
    build_instructions(wb, theme)
    build_dashboard(wb, theme)

    del wb['_temp']

    wb.properties.creator = 'FileSmartCo'
    wb.properties.title = f'Bill & Subscription Tracker - {theme_name}'
    wb.properties.company = 'FileSmartCo'

    return wb


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for name, theme in THEMES.items():
        fname = f'Bill-Subscription-Tracker-{name}.xlsx'
        path = os.path.join(OUTPUT_DIR, fname)
        print(f'Generating {fname}...')
        wb = generate_workbook(name, theme)
        wb.save(path)
        print(f'  \u2713 {path}')
    print('\nDone! All 4 variants generated.')


if __name__ == '__main__':
    main()
