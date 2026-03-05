"""
#009 Mileage Log Template Generator
openpyxl. 4 colour themes. 4 tabs. Native charts.
Same pattern as #006/#007/#008 spreadsheet products.
"""

import os
import sys
from datetime import date, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from colours import THEMES

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection, numbers
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- Constants ----------
LOG_ROWS = 500

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

CATEGORIES = ["Business", "Medical/Moving", "Charity"]

# IRS Standard Mileage Rates
IRS_RATES = {
    2023: {"Business": 0.655, "Medical/Moving": 0.22, "Charity": 0.14},
    2024: {"Business": 0.67, "Medical/Moving": 0.21, "Charity": 0.14},
    2025: {"Business": 0.70, "Medical/Moving": 0.21, "Charity": 0.14},
    2026: {"Business": 0.70, "Medical/Moving": 0.21, "Charity": 0.14},
}

YEARS = [2025, 2026, 2027]

# Sample data
SAMPLE_TRIPS = [
    (date(2025, 1, 6), "Client meeting - Downtown", "Business", "Office", "123 Main St, Downtown", 45230, 45248),
    (date(2025, 1, 8), "Supply pickup - Home Depot", "Business", "Office", "Home Depot, 5th Ave", 45248, 45261),
    (date(2025, 1, 15), "Doctor appointment", "Medical/Moving", "Home", "City Medical Center", 45261, 45276),
    (date(2025, 1, 22), "Deliver donation to Goodwill", "Charity", "Home", "Goodwill, Oak Blvd", 45276, 45284),
    (date(2025, 2, 3), "Job site visit - Lakewood", "Business", "Office", "456 Lake Dr, Lakewood", 45284, 45318),
    (date(2025, 2, 10), "Networking event", "Business", "Home", "Convention Center", 45318, 45340),
    (date(2025, 2, 18), "Tax preparer meeting", "Business", "Office", "H&R Block, Main St", 45340, 45352),
    (date(2025, 3, 5), "Client site inspection", "Business", "Office", "789 Industrial Pkwy", 45352, 45385),
]

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)))

# ---------- Reusable styles ----------

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10, color='333333')
DATA_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='333333')
TOTAL_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')
INPUT_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='0000FF')
TITLE_FONT = Font(name='Arial', size=16, bold=True, color='FFFFFF')
SUBTITLE_FONT = Font(name='Arial', size=12, bold=True, color='333333')
LABEL_FONT = Font(name='Arial', size=10, bold=True, color='333333')
VALUE_FONT = Font(name='Arial', size=14, bold=True, color='1B365D')
SMALL_FONT = Font(name='Arial', size=9, color='718096')
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='718096')

ZEBRA_FILL = PatternFill(fill_type='solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF')

def header_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["primary"])

def accent_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["accent"])

def light_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["light_bg"])

THIN_SIDE = Side(style='thin', color='D9D9D9')
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, bottom=THIN_SIDE, right=THIN_SIDE)
TOTALS_BORDER = Border(top=Side(style='double'), bottom=Side(style='double'),
                        left=THIN_SIDE, right=THIN_SIDE)

def header_border(theme):
    return Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE,
                  bottom=Side(style='medium', color=theme["primary"]))

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
LEFT_INDENT = Alignment(horizontal='left', vertical='center', indent=1)
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ---------- Helpers ----------

def style_header_row(ws, row, start_col, end_col, headers, theme):
    for i, hdr in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = header_border(theme)
    ws.row_dimensions[row].height = 28


def style_data_row(ws, row, start_col, end_col, row_idx):
    ws.row_dimensions[row].height = 22
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.fill = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL


def setup_print(ws, num_cols, last_row, header_rows='1:1', orientation='landscape'):
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_area = f'A1:{get_column_letter(num_cols)}{last_row}'
    ws.print_title_rows = header_rows


# ============================================================
# TAB 1: IRS RATES REFERENCE (built first — referenced by log)
# ============================================================

def build_irs_rates(wb, theme):
    ws = wb.create_sheet("IRS Rates")
    ws.sheet_properties.tabColor = theme["accent"]

    for col, w in enumerate([20, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells('A1:D1')
    title = ws.cell(row=1, column=1, value="IRS Standard Mileage Rates")
    title.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title.fill = header_fill(theme)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 36

    # Rate table header
    headers = ["Year", "Business", "Medical/Moving", "Charity"]
    style_header_row(ws, 3, 1, 4, headers, theme)

    # Rate data
    rate_row = 4
    for i, year in enumerate([2023, 2024, 2025, 2026]):
        row = rate_row + i
        style_data_row(ws, row, 1, 4, i)
        ws.cell(row=row, column=1, value=year).font = DATA_FONT_BOLD
        ws.cell(row=row, column=1).alignment = CENTER
        for j, cat in enumerate(CATEGORIES):
            cell = ws.cell(row=row, column=2 + j, value=IRS_RATES[year][cat])
            cell.number_format = '"$"0.000'
            cell.alignment = CENTER

    # Named ranges for current rates (row 6 = 2025 rates)
    # We'll put the active rates in a clearly marked section
    ws.merge_cells('A9:D9')
    ws.cell(row=9, column=1, value="Active Rates (used for calculations)").font = SUBTITLE_FONT
    ws.cell(row=9, column=1).alignment = LEFT
    ws.row_dimensions[9].height = 28

    active_headers = ["Category", "Rate per Mile"]
    style_header_row(ws, 10, 1, 2, active_headers, theme)

    for i, cat in enumerate(CATEGORIES):
        row = 11 + i
        style_data_row(ws, row, 1, 2, i)
        ws.cell(row=row, column=1, value=cat).font = DATA_FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_INDENT
        cell = ws.cell(row=row, column=2, value=IRS_RATES[2025][cat])
        cell.number_format = '"$"0.000'
        cell.alignment = CENTER
        cell.font = INPUT_FONT_BOLD
        cell.protection = Protection(locked=False)

    # Note about updating rates
    ws.merge_cells('A15:D15')
    ws.cell(row=15, column=1,
            value="To update rates: edit the blue cells above. All calculations will update automatically.").font = NOTE_FONT

    # Category explanations
    ws.merge_cells('A17:D17')
    ws.cell(row=17, column=1, value="Category Explanations").font = SUBTITLE_FONT
    ws.row_dimensions[17].height = 28

    explanations = [
        ("Business", "Driving for business purposes: client visits, job sites, supply runs, business errands. "
                      "Does NOT include commuting from home to your regular workplace."),
        ("Medical/Moving", "Driving to/from medical appointments or for a qualified military move. "
                           "Must be primarily for and essential to medical care."),
        ("Charity", "Driving while performing services for a qualified charitable organization. "
                    "Cannot deduct general commuting to volunteer site."),
    ]

    for i, (cat, desc) in enumerate(explanations):
        row = 18 + i * 2
        ws.cell(row=row, column=1, value=cat).font = DATA_FONT_BOLD
        ws.merge_cells(f'A{row + 1}:D{row + 1}')
        ws.cell(row=row + 1, column=1, value=desc).font = SMALL_FONT
        ws.cell(row=row + 1, column=1).alignment = WRAP
        ws.row_dimensions[row + 1].height = 32

    # Disclaimer
    ws.merge_cells('A25:D25')
    ws.cell(row=25, column=1,
            value="Consult your tax professional for specific deduction eligibility. "
                  "Rates source: IRS.gov").font = NOTE_FONT

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    setup_print(ws, 4, 25, '1:1', 'portrait')
    return ws


# ============================================================
# TAB 2: MILEAGE LOG (main data entry)
# ============================================================

def build_mileage_log(wb, theme):
    ws = wb.create_sheet("Mileage Log")
    ws.sheet_properties.tabColor = theme["primary"]

    # Column widths: A=Date, B=Purpose, C=Category, D=Start Location,
    #   E=End Location, F=Odometer Start, G=Odometer End, H=Miles, I=Deduction, J=Notes
    for col, w in enumerate([12, 28, 16, 20, 20, 14, 14, 12, 14, 22], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    headers = ["Date", "Trip Purpose", "Category", "Start Location",
               "End Location", "Odometer Start", "Odometer End",
               "Miles Driven", "Deduction ($)", "Notes"]
    style_header_row(ws, 1, 1, 10, headers, theme)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:J1'

    # Data validations
    dv_cat = DataValidation(
        type='list', formula1='"' + ','.join(CATEGORIES) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Category', error='Select: Business, Medical/Moving, or Charity')
    ws.add_data_validation(dv_cat)

    data_start = 2
    for i in range(LOG_ROWS):
        row = data_start + i
        style_data_row(ws, row, 1, 10, i)

        # A: Date (unlocked)
        cell = ws.cell(row=row, column=1)
        cell.font = INPUT_FONT
        cell.number_format = 'MM/DD/YYYY'
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)

        # B: Trip Purpose (unlocked)
        cell = ws.cell(row=row, column=2)
        cell.font = INPUT_FONT
        cell.alignment = LEFT_INDENT
        cell.protection = Protection(locked=False)

        # C: Category (unlocked, dropdown)
        cell = ws.cell(row=row, column=3)
        cell.font = INPUT_FONT
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)
        dv_cat.add(cell)

        # D: Start Location (unlocked)
        cell = ws.cell(row=row, column=4)
        cell.font = INPUT_FONT
        cell.alignment = LEFT
        cell.protection = Protection(locked=False)

        # E: End Location (unlocked)
        cell = ws.cell(row=row, column=5)
        cell.font = INPUT_FONT
        cell.alignment = LEFT
        cell.protection = Protection(locked=False)

        # F: Odometer Start (unlocked)
        cell = ws.cell(row=row, column=6)
        cell.font = INPUT_FONT
        cell.number_format = '#,##0'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # G: Odometer End (unlocked)
        cell = ws.cell(row=row, column=7)
        cell.font = INPUT_FONT
        cell.number_format = '#,##0'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # H: Miles Driven (auto-calc: End - Start, only if both filled)
        cell = ws.cell(row=row, column=8)
        cell.value = f'=IF(AND(G{row}<>"",F{row}<>""),G{row}-F{row},"")'
        cell.font = DATA_FONT_BOLD
        cell.number_format = '#,##0.0'
        cell.alignment = CENTER

        # I: Deduction (auto-calc: Miles × Rate from IRS Rates tab)
        # VLOOKUP category in IRS Rates!A11:B13 to get rate, multiply by miles
        cell = ws.cell(row=row, column=9)
        cell.value = (f'=IF(AND(H{row}<>"",C{row}<>""),'
                      f'H{row}*VLOOKUP(C{row},\'IRS Rates\'!$A$11:$B$13,2,FALSE),"")')
        cell.font = DATA_FONT_BOLD
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT

        # J: Notes (unlocked)
        cell = ws.cell(row=row, column=10)
        cell.font = INPUT_FONT
        cell.alignment = LEFT
        cell.protection = Protection(locked=False)

    # Totals row
    totals_row = data_start + LOG_ROWS
    ws.row_dimensions[totals_row].height = 28
    ws.merge_cells(f'A{totals_row}:G{totals_row}')
    cell = ws.cell(row=totals_row, column=1, value="TOTALS")
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.alignment = CENTER
    cell.border = TOTALS_BORDER
    for col in range(2, 8):
        ws.cell(row=totals_row, column=col).fill = header_fill(theme)
        ws.cell(row=totals_row, column=col).border = TOTALS_BORDER

    # Total miles
    cell = ws.cell(row=totals_row, column=8)
    cell.value = f'=SUMPRODUCT((H2:H{data_start + LOG_ROWS - 1}<>"")*H2:H{data_start + LOG_ROWS - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '#,##0.0'
    cell.alignment = CENTER
    cell.border = TOTALS_BORDER

    # Total deduction
    cell = ws.cell(row=totals_row, column=9)
    cell.value = f'=SUMPRODUCT((I2:I{data_start + LOG_ROWS - 1}<>"")*I2:I{data_start + LOG_ROWS - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    ws.cell(row=totals_row, column=10).fill = header_fill(theme)
    ws.cell(row=totals_row, column=10).border = TOTALS_BORDER

    # Fill sample data
    for idx, (dt, purpose, cat, start_loc, end_loc, odo_start, odo_end) in enumerate(SAMPLE_TRIPS):
        row = data_start + idx
        ws.cell(row=row, column=1, value=dt)
        ws.cell(row=row, column=2, value=purpose)
        ws.cell(row=row, column=3, value=cat)
        ws.cell(row=row, column=4, value=start_loc)
        ws.cell(row=row, column=5, value=end_loc)
        ws.cell(row=row, column=6, value=odo_start)
        ws.cell(row=row, column=7, value=odo_end)

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    setup_print(ws, 10, totals_row, '1:1', 'landscape')
    return ws


# ============================================================
# TAB 3: MONTHLY SUMMARY
# ============================================================

def build_monthly_summary(wb, theme):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = theme["highlight"]

    for col, w in enumerate([16, 16, 18, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells('A1:F1')
    title = ws.cell(row=1, column=1, value="Monthly Mileage Summary")
    title.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title.fill = header_fill(theme)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 36

    headers = ["Month", "Business Miles", "Medical/Moving Miles", "Charity Miles",
               "Total Miles", "Total Deduction"]
    style_header_row(ws, 3, 1, 6, headers, theme)

    log_range_end = 501  # rows 2-501 in Mileage Log

    for i, month in enumerate(MONTHS):
        row = 4 + i
        month_num = i + 1
        style_data_row(ws, row, 1, 6, i)

        ws.cell(row=row, column=1, value=month).font = DATA_FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_INDENT

        # Business miles: SUMPRODUCT to filter by month and category
        ws.cell(row=row, column=2).value = (
            f'=SUMPRODUCT((MONTH(\'Mileage Log\'!A2:A{log_range_end})={month_num})'
            f'*(\'Mileage Log\'!C2:C{log_range_end}="Business")'
            f'*(\'Mileage Log\'!H2:H{log_range_end}<>"")'
            f'*\'Mileage Log\'!H2:H{log_range_end})')
        ws.cell(row=row, column=2).number_format = '#,##0.0'
        ws.cell(row=row, column=2).alignment = RIGHT

        # Medical/Moving miles
        ws.cell(row=row, column=3).value = (
            f'=SUMPRODUCT((MONTH(\'Mileage Log\'!A2:A{log_range_end})={month_num})'
            f'*(\'Mileage Log\'!C2:C{log_range_end}="Medical/Moving")'
            f'*(\'Mileage Log\'!H2:H{log_range_end}<>"")'
            f'*\'Mileage Log\'!H2:H{log_range_end})')
        ws.cell(row=row, column=3).number_format = '#,##0.0'
        ws.cell(row=row, column=3).alignment = RIGHT

        # Charity miles
        ws.cell(row=row, column=4).value = (
            f'=SUMPRODUCT((MONTH(\'Mileage Log\'!A2:A{log_range_end})={month_num})'
            f'*(\'Mileage Log\'!C2:C{log_range_end}="Charity")'
            f'*(\'Mileage Log\'!H2:H{log_range_end}<>"")'
            f'*\'Mileage Log\'!H2:H{log_range_end})')
        ws.cell(row=row, column=4).number_format = '#,##0.0'
        ws.cell(row=row, column=4).alignment = RIGHT

        # Total miles
        ws.cell(row=row, column=5).value = f'=SUM(B{row}:D{row})'
        ws.cell(row=row, column=5).number_format = '#,##0.0'
        ws.cell(row=row, column=5).font = DATA_FONT_BOLD
        ws.cell(row=row, column=5).alignment = RIGHT

        # Total deduction
        ws.cell(row=row, column=6).value = (
            f'=SUMPRODUCT((MONTH(\'Mileage Log\'!A2:A{log_range_end})={month_num})'
            f'*(\'Mileage Log\'!I2:I{log_range_end}<>"")'
            f'*\'Mileage Log\'!I2:I{log_range_end})')
        ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=6).font = DATA_FONT_BOLD
        ws.cell(row=row, column=6).alignment = RIGHT

    # Annual totals row
    totals_row = 16
    ws.row_dimensions[totals_row].height = 28
    ws.cell(row=totals_row, column=1, value="ANNUAL TOTAL").font = TOTAL_FONT
    ws.cell(row=totals_row, column=1).fill = header_fill(theme)
    ws.cell(row=totals_row, column=1).alignment = CENTER
    ws.cell(row=totals_row, column=1).border = TOTALS_BORDER

    for col in range(2, 7):
        cell = ws.cell(row=totals_row, column=col)
        cell.value = f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}15)'
        cell.font = TOTAL_FONT
        cell.fill = header_fill(theme)
        cell.alignment = RIGHT
        cell.border = TOTALS_BORDER
        if col == 6:
            cell.number_format = '"$"#,##0.00'
        else:
            cell.number_format = '#,##0.0'

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    setup_print(ws, 6, 16, '1:3', 'landscape')
    return ws


# ============================================================
# TAB 4: DASHBOARD (built last — references other tabs)
# ============================================================

def build_dashboard(wb, theme, theme_name):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = theme["primary"]

    for col, w in enumerate([3, 20, 18, 18, 18, 18, 3], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title bar
    ws.merge_cells('A1:G1')
    title = ws.cell(row=1, column=1, value="MILEAGE LOG")
    title.font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
    title.fill = header_fill(theme)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 48

    ws.merge_cells('A2:G2')
    sub = ws.cell(row=2, column=1, value="Business Mileage Tracker & IRS Deduction Calculator")
    sub.font = Font(name='Arial', size=11, color=theme["muted_text"])
    sub.fill = light_fill(theme)
    sub.alignment = CENTER
    ws.row_dimensions[2].height = 28

    # Year selector
    ws.cell(row=4, column=2, value="Tax Year:").font = LABEL_FONT
    ws.cell(row=4, column=2).alignment = RIGHT
    year_cell = ws.cell(row=4, column=3, value=2025)
    year_cell.font = Font(name='Arial', size=14, bold=True, color='0000FF')
    year_cell.alignment = CENTER
    year_cell.protection = Protection(locked=False)
    year_cell.border = Border(bottom=Side(style='medium', color=theme["primary"]))

    dv_year = DataValidation(
        type='list', formula1='"2025,2026,2027"',
        allow_blank=False, showErrorMessage=True,
        errorTitle='Invalid Year', error='Select 2025, 2026, or 2027')
    ws.add_data_validation(dv_year)
    dv_year.add(year_cell)

    ws.row_dimensions[4].height = 30

    # --- KPI Cards Row ---
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 36

    log_range_end = 501
    year_ref = '$C$4'

    # Card 1: Total Miles
    ws.cell(row=6, column=2, value="Total Miles").font = LABEL_FONT
    ws.cell(row=6, column=2).alignment = CENTER
    cell = ws.cell(row=7, column=2)
    cell.value = (f'=SUMPRODUCT((YEAR(\'Mileage Log\'!A2:A{log_range_end})={year_ref})'
                  f'*(\'Mileage Log\'!H2:H{log_range_end}<>"")'
                  f'*\'Mileage Log\'!H2:H{log_range_end})')
    cell.font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    cell.number_format = '#,##0.0'
    cell.alignment = CENTER

    # Card 2: Total Deduction
    ws.cell(row=6, column=3, value="Total Deduction").font = LABEL_FONT
    ws.cell(row=6, column=3).alignment = CENTER
    cell = ws.cell(row=7, column=3)
    cell.value = (f'=SUMPRODUCT((YEAR(\'Mileage Log\'!A2:A{log_range_end})={year_ref})'
                  f'*(\'Mileage Log\'!I2:I{log_range_end}<>"")'
                  f'*\'Mileage Log\'!I2:I{log_range_end})')
    cell.font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # Card 3: Total Trips
    ws.cell(row=6, column=4, value="Total Trips").font = LABEL_FONT
    ws.cell(row=6, column=4).alignment = CENTER
    cell = ws.cell(row=7, column=4)
    cell.value = (f'=SUMPRODUCT((YEAR(\'Mileage Log\'!A2:A{log_range_end})={year_ref})'
                  f'*(\'Mileage Log\'!A2:A{log_range_end}<>""))')
    cell.font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    cell.number_format = '#,##0'
    cell.alignment = CENTER

    # --- Miles by Category ---
    ws.row_dimensions[9].height = 28
    ws.merge_cells('B9:F9')
    ws.cell(row=9, column=2, value="Miles by Category").font = SUBTITLE_FONT
    ws.cell(row=9, column=2).alignment = LEFT

    cat_headers = ["Category", "Miles", "Deduction", "% of Total"]
    style_header_row(ws, 10, 2, 5, cat_headers, theme)

    for i, cat in enumerate(CATEGORIES):
        row = 11 + i
        style_data_row(ws, row, 2, 5, i)
        ws.cell(row=row, column=2, value=cat).font = DATA_FONT_BOLD
        ws.cell(row=row, column=2).alignment = LEFT_INDENT

        # Miles for category
        ws.cell(row=row, column=3).value = (
            f'=SUMPRODUCT((YEAR(\'Mileage Log\'!A2:A{log_range_end})={year_ref})'
            f'*(\'Mileage Log\'!C2:C{log_range_end}="{cat}")'
            f'*(\'Mileage Log\'!H2:H{log_range_end}<>"")'
            f'*\'Mileage Log\'!H2:H{log_range_end})')
        ws.cell(row=row, column=3).number_format = '#,##0.0'
        ws.cell(row=row, column=3).alignment = RIGHT

        # Deduction for category
        ws.cell(row=row, column=4).value = (
            f'=SUMPRODUCT((YEAR(\'Mileage Log\'!A2:A{log_range_end})={year_ref})'
            f'*(\'Mileage Log\'!C2:C{log_range_end}="{cat}")'
            f'*(\'Mileage Log\'!I2:I{log_range_end}<>"")'
            f'*\'Mileage Log\'!I2:I{log_range_end})')
        ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4).alignment = RIGHT

        # % of total
        ws.cell(row=row, column=5).value = f'=IF($C$7=0,0,C{row}/$C$7)'
        ws.cell(row=row, column=5).number_format = '0.0%'
        ws.cell(row=row, column=5).alignment = CENTER

    # --- IRS Rates Display ---
    ws.row_dimensions[15].height = 28
    ws.merge_cells('B15:F15')
    ws.cell(row=15, column=2, value="Current IRS Standard Mileage Rates (2025)").font = SUBTITLE_FONT

    rate_headers = ["Category", "Rate per Mile"]
    style_header_row(ws, 16, 2, 3, rate_headers, theme)

    for i, cat in enumerate(CATEGORIES):
        row = 17 + i
        style_data_row(ws, row, 2, 3, i)
        ws.cell(row=row, column=2, value=cat).font = DATA_FONT_BOLD
        ws.cell(row=row, column=2).alignment = LEFT_INDENT
        ws.cell(row=row, column=3).value = f'=\'IRS Rates\'!B{11 + i}'
        ws.cell(row=row, column=3).number_format = '"$"0.000"/mile"'
        ws.cell(row=row, column=3).alignment = CENTER

    # --- Miles by Month Bar Chart ---
    ws.row_dimensions[21].height = 28
    ws.merge_cells('B21:F21')
    ws.cell(row=21, column=2, value="Miles by Month").font = SUBTITLE_FONT

    # Chart data table (hidden area, used by chart)
    chart_data_row = 22
    ws.cell(row=chart_data_row, column=2, value="Month").font = SMALL_FONT
    ws.cell(row=chart_data_row, column=3, value="Miles").font = SMALL_FONT

    for i, month in enumerate(MONTHS_SHORT):
        row = chart_data_row + 1 + i
        month_num = i + 1
        ws.cell(row=row, column=2, value=month).font = SMALL_FONT
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3).value = (
            f'=SUMPRODUCT((YEAR(\'Mileage Log\'!A2:A{log_range_end})={year_ref})'
            f'*(MONTH(\'Mileage Log\'!A2:A{log_range_end})={month_num})'
            f'*(\'Mileage Log\'!H2:H{log_range_end}<>"")'
            f'*\'Mileage Log\'!H2:H{log_range_end})')
        ws.cell(row=row, column=3).number_format = '#,##0.0'

    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = None
    chart.y_axis.title = "Miles"
    chart.x_axis.title = None
    chart.y_axis.numFmt = '#,##0'

    data = Reference(ws, min_col=3, min_row=chart_data_row, max_row=chart_data_row + 12)
    cats = Reference(ws, min_col=2, min_row=chart_data_row + 1, max_row=chart_data_row + 12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 28
    chart.height = 12

    # Theme color for bars
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    series = chart.series[0]
    series.graphicalProperties.solidFill = theme["primary"]

    ws.add_chart(chart, "B36")

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    return ws


# ============================================================
# MAIN GENERATOR
# ============================================================

def generate(theme_name, theme):
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Build tabs in order (IRS Rates first — referenced by Mileage Log)
    build_irs_rates(wb, theme)
    build_mileage_log(wb, theme)
    build_monthly_summary(wb, theme)
    build_dashboard(wb, theme, theme_name)

    # Reorder: Dashboard first
    wb.move_sheet("Dashboard", offset=-3)

    # Set Dashboard as active sheet
    wb.active = wb.sheetnames.index("Dashboard")

    # Save
    safe_name = theme_name.replace(" ", "-")
    filename = f"Mileage-Log-{safe_name}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"  Generated: {filepath}")
    return filepath


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Generating #009 Mileage Log Template...")
    print(f"Output: {OUTPUT_DIR}")
    print()

    files = []
    for theme_name, theme in THEMES.items():
        f = generate(theme_name, theme)
        files.append(f)

    print()
    print(f"Done! {len(files)} files generated.")
    for f in files:
        print(f"  {os.path.basename(f)}")


if __name__ == "__main__":
    main()
