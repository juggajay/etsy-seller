"""
#010 Tax Deduction Tracker Generator
openpyxl. 4 colour themes. 5 tabs. Native charts.
Same pattern as #006/#007/#008/#009 spreadsheet products.
"""

import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from colours import THEMES

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- Constants ----------
LOG_ROWS = 500

MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

CATEGORIES = [
    "Advertising & Marketing",
    "Car & Vehicle Expenses",
    "Contract Labour",
    "Home Office",
    "Insurance (Business)",
    "Interest (Business Loans)",
    "Legal & Professional Services",
    "Office Supplies & Equipment",
    "Rent (Business Property)",
    "Repairs & Maintenance",
    "Software & Subscriptions",
    "Telephone & Internet",
    "Travel & Lodging",
    "Meals (50% deductible)",
    "Utilities",
    "Education & Training",
    "Bank & Processing Fees",
    "Other / Miscellaneous",
]
NUM_CATS = len(CATEGORIES)

PAYMENT_METHODS = ["Cash", "Credit Card", "Bank Transfer", "Check", "Other"]
RECEIPT_OPTIONS = ["Yes", "No", "Digital"]

# Sample data
SAMPLE_DEDUCTIONS = [
    (date(2025, 1, 5), "Google Ads campaign", "Advertising & Marketing", "Google", 450.00, "Credit Card", "Digital"),
    (date(2025, 1, 12), "QuickBooks subscription", "Software & Subscriptions", "Intuit", 35.00, "Credit Card", "Digital"),
    (date(2025, 1, 18), "Office desk and chair", "Office Supplies & Equipment", "IKEA", 389.99, "Credit Card", "Yes"),
    (date(2025, 1, 25), "Business lunch with client", "Meals (50% deductible)", "Olive Garden", 78.50, "Credit Card", "Yes"),
    (date(2025, 2, 1), "Monthly rent", "Rent (Business Property)", "ABC Properties", 1200.00, "Bank Transfer", "Yes"),
    (date(2025, 2, 8), "Internet service", "Telephone & Internet", "Comcast", 89.99, "Bank Transfer", "Digital"),
    (date(2025, 2, 15), "Tax prep consultation", "Legal & Professional Services", "H&R Block", 250.00, "Check", "Yes"),
    (date(2025, 2, 22), "Home office deduction", "Home Office", "Self", 175.00, "Bank Transfer", "No"),
    (date(2025, 3, 1), "Monthly rent", "Rent (Business Property)", "ABC Properties", 1200.00, "Bank Transfer", "Yes"),
    (date(2025, 3, 5), "Online course - Marketing", "Education & Training", "Coursera", 49.99, "Credit Card", "Digital"),
    (date(2025, 3, 12), "Printer paper and ink", "Office Supplies & Equipment", "Staples", 67.45, "Cash", "Yes"),
    (date(2025, 3, 20), "Stripe processing fees", "Bank & Processing Fees", "Stripe", 42.30, "Bank Transfer", "Digital"),
]

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)))

# ---------- Reusable styles ----------

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10, color='333333')
DATA_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='333333')
TOTAL_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')
INPUT_FONT_BOLD = Font(name='Arial', size=10, bold=True, color='0000FF')
TITLE_FONT = Font(name='Arial', size=16, bold=True, color='FFFFFF')
SUBTITLE_FONT = Font(name='Arial', size=12, bold=True, color='333333')
LABEL_FONT = Font(name='Arial', size=10, bold=True, color='333333')
SMALL_FONT = Font(name='Arial', size=9, color='718096')
NOTE_FONT = Font(name='Arial', size=9, italic=True, color='718096')

ZEBRA_FILL = PatternFill(fill_type='solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF')

def header_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["primary"])

def accent_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["accent"])

def light_fill(theme):
    return PatternFill(fill_type='solid', fgColor=theme["light_bg"])

THIN_SIDE = Side(style='thin', color='D9D9D9')
THIN_BORDER = Border(top=THIN_SIDE, left=THIN_SIDE, bottom=THIN_SIDE, right=THIN_SIDE)
TOTALS_BORDER = Border(top=Side(style='double'), bottom=Side(style='double'),
                        left=THIN_SIDE, right=THIN_SIDE)

def header_border(theme):
    return Border(top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE,
                  bottom=Side(style='medium', color=theme["primary"]))

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
LEFT_INDENT = Alignment(horizontal='left', vertical='center', indent=1)
RIGHT = Alignment(horizontal='right', vertical='center')
WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ---------- Helpers ----------

def style_header_row(ws, row, start_col, end_col, headers, theme):
    for i, hdr in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = header_fill(theme)
        cell.alignment = CENTER
        cell.border = header_border(theme)
    ws.row_dimensions[row].height = 28


def style_data_row(ws, row, start_col, end_col, row_idx):
    ws.row_dimensions[row].height = 22
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.fill = ZEBRA_FILL if row_idx % 2 == 0 else WHITE_FILL


def setup_print(ws, num_cols, last_row, header_rows='1:1', orientation='landscape',
                fit_height=0):
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_area = f'A1:{get_column_letter(num_cols)}{last_row}'
    ws.print_title_rows = header_rows


# ============================================================
# TAB 1: DEDUCTION LOG (main data entry)
# ============================================================

def build_deduction_log(wb, theme):
    ws = wb.create_sheet("Deduction Log")
    ws.sheet_properties.tabColor = theme["primary"]

    # Columns: A=Date, B=Description, C=Category, D=Vendor, E=Amount,
    #   F=Payment Method, G=Receipt?, H=Notes
    for col, w in enumerate([12, 30, 26, 20, 14, 16, 12, 24], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    headers = ["Date", "Description", "Category", "Vendor / Payee",
               "Amount ($)", "Payment Method", "Receipt?", "Notes"]
    style_header_row(ws, 1, 1, 8, headers, theme)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:H1'

    # Data validations
    dv_cat = DataValidation(
        type='list', formula1='"' + ','.join(CATEGORIES) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Category', error='Select a deduction category from the dropdown')
    dv_pay = DataValidation(
        type='list', formula1='"' + ','.join(PAYMENT_METHODS) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid Payment Method', error='Select a payment method from the dropdown')
    dv_receipt = DataValidation(
        type='list', formula1='"' + ','.join(RECEIPT_OPTIONS) + '"',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Invalid', error='Select Yes, No, or Digital')
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_pay)
    ws.add_data_validation(dv_receipt)

    data_start = 2
    for i in range(LOG_ROWS):
        row = data_start + i
        style_data_row(ws, row, 1, 8, i)

        # A: Date (unlocked)
        cell = ws.cell(row=row, column=1)
        cell.font = INPUT_FONT
        cell.number_format = 'MM/DD/YYYY'
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)

        # B: Description (unlocked)
        cell = ws.cell(row=row, column=2)
        cell.font = INPUT_FONT
        cell.alignment = LEFT_INDENT
        cell.protection = Protection(locked=False)

        # C: Category (unlocked, dropdown)
        cell = ws.cell(row=row, column=3)
        cell.font = INPUT_FONT
        cell.alignment = LEFT
        cell.protection = Protection(locked=False)
        dv_cat.add(cell)

        # D: Vendor (unlocked)
        cell = ws.cell(row=row, column=4)
        cell.font = INPUT_FONT
        cell.alignment = LEFT
        cell.protection = Protection(locked=False)

        # E: Amount (unlocked)
        cell = ws.cell(row=row, column=5)
        cell.font = INPUT_FONT
        cell.number_format = '"$"#,##0.00'
        cell.alignment = RIGHT
        cell.protection = Protection(locked=False)

        # F: Payment Method (unlocked, dropdown)
        cell = ws.cell(row=row, column=6)
        cell.font = INPUT_FONT
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)
        dv_pay.add(cell)

        # G: Receipt? (unlocked, dropdown)
        cell = ws.cell(row=row, column=7)
        cell.font = INPUT_FONT
        cell.alignment = CENTER
        cell.protection = Protection(locked=False)
        dv_receipt.add(cell)

        # H: Notes (unlocked)
        cell = ws.cell(row=row, column=8)
        cell.font = INPUT_FONT
        cell.alignment = LEFT
        cell.protection = Protection(locked=False)

    # Totals row
    totals_row = data_start + LOG_ROWS
    ws.row_dimensions[totals_row].height = 28
    ws.merge_cells(f'A{totals_row}:D{totals_row}')
    cell = ws.cell(row=totals_row, column=1, value="TOTAL DEDUCTIONS")
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.alignment = CENTER
    cell.border = TOTALS_BORDER
    for col in range(2, 5):
        ws.cell(row=totals_row, column=col).fill = header_fill(theme)
        ws.cell(row=totals_row, column=col).border = TOTALS_BORDER

    cell = ws.cell(row=totals_row, column=5)
    cell.value = f'=SUM(E2:E{data_start + LOG_ROWS - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    for col in range(6, 9):
        ws.cell(row=totals_row, column=col).fill = header_fill(theme)
        ws.cell(row=totals_row, column=col).border = TOTALS_BORDER

    # Fill sample data
    for idx, (dt, desc, cat, vendor, amt, pay, receipt) in enumerate(SAMPLE_DEDUCTIONS):
        row = data_start + idx
        ws.cell(row=row, column=1, value=dt)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=cat)
        ws.cell(row=row, column=4, value=vendor)
        ws.cell(row=row, column=5, value=amt)
        ws.cell(row=row, column=6, value=pay)
        ws.cell(row=row, column=7, value=receipt)

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    setup_print(ws, 8, totals_row, '1:1', 'landscape')
    return ws


# ============================================================
# TAB 2: CATEGORY SUMMARY
# ============================================================

def build_category_summary(wb, theme):
    ws = wb.create_sheet("Category Summary")
    ws.sheet_properties.tabColor = theme["accent"]

    for col, w in enumerate([30, 16, 14, 14, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells('A1:E1')
    title = ws.cell(row=1, column=1, value="Tax Deduction Summary by Category")
    title.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title.fill = header_fill(theme)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 36

    headers = ["Category", "Total Amount", "# of Entries", "% of Total", "Deductible Amount"]
    style_header_row(ws, 3, 1, 5, headers, theme)

    log_range_end = 501  # rows 2-501 in Deduction Log
    grand_total_row = 3 + NUM_CATS + 1

    for i, cat in enumerate(CATEGORIES):
        row = 4 + i
        style_data_row(ws, row, 1, 5, i)

        ws.cell(row=row, column=1, value=cat).font = DATA_FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_INDENT

        # Total amount (SUMIF)
        ws.cell(row=row, column=2).value = (
            f'=SUMPRODUCT((\'Deduction Log\'!C2:C{log_range_end}="{cat}")'
            f'*\'Deduction Log\'!E2:E{log_range_end})')
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=2).alignment = RIGHT

        # Count of entries
        ws.cell(row=row, column=3).value = (
            f'=COUNTIF(\'Deduction Log\'!C2:C{log_range_end},"{cat}")')
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).alignment = CENTER

        # % of total
        ws.cell(row=row, column=4).value = (
            f'=IF(B{grand_total_row}=0,0,B{row}/B{grand_total_row})')
        ws.cell(row=row, column=4).number_format = '0.0%'
        ws.cell(row=row, column=4).alignment = CENTER

        # Deductible amount (Meals = 50%, everything else = 100%)
        if cat == "Meals (50% deductible)":
            ws.cell(row=row, column=5).value = f'=B{row}*0.5'
        else:
            ws.cell(row=row, column=5).value = f'=B{row}'
        ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=5).alignment = RIGHT
        ws.cell(row=row, column=5).font = DATA_FONT_BOLD

    # Grand total row
    ws.row_dimensions[grand_total_row].height = 28
    ws.cell(row=grand_total_row, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=grand_total_row, column=1).fill = header_fill(theme)
    ws.cell(row=grand_total_row, column=1).alignment = CENTER
    ws.cell(row=grand_total_row, column=1).border = TOTALS_BORDER

    # Total amount
    cell = ws.cell(row=grand_total_row, column=2)
    cell.value = f'=SUM(B4:B{grand_total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Total entries
    cell = ws.cell(row=grand_total_row, column=3)
    cell.value = f'=SUM(C4:C{grand_total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '#,##0'
    cell.alignment = CENTER
    cell.border = TOTALS_BORDER

    # % (100%)
    cell = ws.cell(row=grand_total_row, column=4)
    cell.value = 1
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '0.0%'
    cell.alignment = CENTER
    cell.border = TOTALS_BORDER

    # Total deductible
    cell = ws.cell(row=grand_total_row, column=5)
    cell.value = f'=SUM(E4:E{grand_total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Note about meals
    note_row = grand_total_row + 2
    ws.merge_cells(f'A{note_row}:E{note_row}')
    ws.cell(row=note_row, column=1,
            value='Note: Meals are only 50% deductible per IRS rules. '
                  'The "Deductible Amount" column reflects this adjustment.').font = NOTE_FONT

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    setup_print(ws, 5, note_row, '1:3', 'portrait')
    return ws


# ============================================================
# TAB 3: ANNUAL TAX SUMMARY (print-ready, 1 page)
# ============================================================

def build_annual_summary(wb, theme):
    ws = wb.create_sheet("Annual Tax Summary")
    ws.sheet_properties.tabColor = theme["highlight"]

    for col, w in enumerate([30, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells('A1:C1')
    title = ws.cell(row=1, column=1, value="ANNUAL TAX DEDUCTION SUMMARY")
    title.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title.fill = header_fill(theme)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 36

    # Business info fields
    info_fields = [
        ("Business / Individual Name:", 3),
        ("EIN / SSN:", 4),
        ("Tax Year:", 5),
        ("Prepared By:", 6),
    ]

    for label, row in info_fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).alignment = RIGHT
        cell = ws.cell(row=row, column=2)
        cell.font = INPUT_FONT_BOLD
        cell.alignment = LEFT
        cell.protection = Protection(locked=False)
        cell.border = Border(bottom=Side(style='thin', color=theme["primary"]))
        ws.merge_cells(f'B{row}:C{row}')
        ws.row_dimensions[row].height = 24

    # Pre-fill tax year
    ws.cell(row=5, column=2, value=2025)

    # Category totals table
    ws.row_dimensions[8].height = 28
    ws.merge_cells('A8:C8')
    ws.cell(row=8, column=1, value="Deduction Breakdown").font = SUBTITLE_FONT

    headers = ["Category", "Total Spent", "Deductible Amount"]
    style_header_row(ws, 9, 1, 3, headers, theme)

    for i, cat in enumerate(CATEGORIES):
        row = 10 + i
        style_data_row(ws, row, 1, 3, i)

        ws.cell(row=row, column=1, value=cat).font = DATA_FONT
        ws.cell(row=row, column=1).alignment = LEFT_INDENT

        # Total from Category Summary
        cat_sum_row = 4 + i
        ws.cell(row=row, column=2).value = f'=\'Category Summary\'!B{cat_sum_row}'
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=2).alignment = RIGHT

        ws.cell(row=row, column=3).value = f'=\'Category Summary\'!E{cat_sum_row}'
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=3).alignment = RIGHT

    # Total row
    total_row = 10 + NUM_CATS
    ws.row_dimensions[total_row].height = 28

    ws.cell(row=total_row, column=1, value="TOTAL DEDUCTIONS").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = header_fill(theme)
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=1).border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=2)
    cell.value = f'=SUM(B10:B{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    cell = ws.cell(row=total_row, column=3)
    cell.value = f'=SUM(C10:C{total_row - 1})'
    cell.font = TOTAL_FONT
    cell.fill = header_fill(theme)
    cell.number_format = '"$"#,##0.00'
    cell.alignment = RIGHT
    cell.border = TOTALS_BORDER

    # Signature section
    sig_row = total_row + 3
    ws.cell(row=sig_row, column=1, value="Signature:").font = LABEL_FONT
    ws.cell(row=sig_row, column=1).alignment = RIGHT
    ws.merge_cells(f'B{sig_row}:C{sig_row}')
    ws.cell(row=sig_row, column=2).border = Border(
        bottom=Side(style='thin', color='333333'))
    ws.row_dimensions[sig_row].height = 30

    ws.cell(row=sig_row + 1, column=1, value="Date:").font = LABEL_FONT
    ws.cell(row=sig_row + 1, column=1).alignment = RIGHT
    ws.merge_cells(f'B{sig_row + 1}:C{sig_row + 1}')
    ws.cell(row=sig_row + 1, column=2).border = Border(
        bottom=Side(style='thin', color='333333'))
    ws.cell(row=sig_row + 1, column=2).protection = Protection(locked=False)

    # Disclaimer
    disc_row = sig_row + 3
    ws.merge_cells(f'A{disc_row}:C{disc_row}')
    ws.cell(row=disc_row, column=1,
            value="This summary is for reference only. Consult a qualified tax "
                  "professional for specific deduction eligibility and filing requirements.").font = NOTE_FONT
    ws.cell(row=disc_row, column=1).alignment = WRAP
    ws.row_dimensions[disc_row].height = 28

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    # Print setup: FIT TO 1 PAGE
    setup_print(ws, 3, disc_row, '1:1', 'portrait', fit_height=1)
    return ws


# ============================================================
# TAB 4: DEDUCTION GUIDE
# ============================================================

def build_deduction_guide(wb, theme):
    ws = wb.create_sheet("Deduction Guide")
    ws.sheet_properties.tabColor = theme["card_border"]

    for col, w in enumerate([30, 50], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells('A1:B1')
    title = ws.cell(row=1, column=1, value="Tax Deduction Guide & Tips")
    title.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title.fill = header_fill(theme)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 36

    # Category explanations
    ws.merge_cells('A3:B3')
    ws.cell(row=3, column=1, value="Category Explanations").font = SUBTITLE_FONT
    ws.row_dimensions[3].height = 28

    explanations = [
        ("Advertising & Marketing", "Online ads, business cards, website costs, social media ads, signage"),
        ("Car & Vehicle Expenses", "Gas, maintenance, insurance, depreciation OR standard mileage rate"),
        ("Contract Labour", "Payments to freelancers, subcontractors (1099 workers)"),
        ("Home Office", "Dedicated workspace: % of rent/mortgage, utilities, insurance"),
        ("Insurance (Business)", "Business liability, professional indemnity, workers comp"),
        ("Interest (Business Loans)", "Interest on business loans, business credit cards"),
        ("Legal & Professional Services", "Attorney fees, accountant fees, tax preparation"),
        ("Office Supplies & Equipment", "Computer, printer, paper, pens, software, furniture"),
        ("Rent (Business Property)", "Office rent, warehouse, storage unit for business"),
        ("Repairs & Maintenance", "Equipment repair, office maintenance, vehicle repair"),
        ("Software & Subscriptions", "SaaS tools, cloud storage, professional memberships"),
        ("Telephone & Internet", "Business phone, internet service (business % if shared)"),
        ("Travel & Lodging", "Flights, hotels, car rental for business travel"),
        ("Meals (50% deductible)", "Business meals with clients/partners — only 50% deductible"),
        ("Utilities", "Electric, gas, water for business property"),
        ("Education & Training", "Courses, certifications, workshops, books for business"),
        ("Bank & Processing Fees", "Bank fees, credit card processing, PayPal/Stripe fees"),
        ("Other / Miscellaneous", "Any legitimate business expense not in the above categories"),
    ]

    row = 4
    for cat, desc in explanations:
        ws.cell(row=row, column=1, value=cat).font = DATA_FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT_INDENT
        ws.cell(row=row, column=2, value=desc).font = DATA_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        ws.row_dimensions[row].height = 22
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=1).fill = ZEBRA_FILL if (row - 4) % 2 == 0 else WHITE_FILL
        ws.cell(row=row, column=2).fill = ZEBRA_FILL if (row - 4) % 2 == 0 else WHITE_FILL
        row += 1

    # Commonly missed deductions
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Commonly Missed Deductions").font = SUBTITLE_FONT
    ws.row_dimensions[row].height = 28
    row += 1

    missed = [
        "Home office deduction (simplified: $5/sq ft, up to 300 sq ft = $1,500)",
        "Cell phone bill (business use percentage)",
        "Internet service (business use percentage)",
        "Professional development courses and books",
        "Business use of personal vehicle (track mileage!)",
        "Health insurance premiums (self-employed)",
        "Retirement plan contributions (SEP IRA, Solo 401k)",
        "Business bank account fees and credit card annual fees",
        "Professional association memberships and dues",
        "Software and app subscriptions used for business",
    ]

    for item in missed:
        ws.cell(row=row, column=1, value="\u2022").font = DATA_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=item).font = DATA_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        ws.row_dimensions[row].height = 20
        row += 1

    # Record-keeping tips
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="Record-Keeping Tips").font = SUBTITLE_FONT
    ws.row_dimensions[row].height = 28
    row += 1

    tips = [
        "Keep ALL receipts — photograph paper receipts and store digitally",
        "Record expenses as they happen, not at year-end",
        "Separate business and personal bank accounts/credit cards",
        "Track mileage with an app or dedicated log (see our Mileage Log template!)",
        "Save invoices, contracts, and payment confirmations",
        "Back up digital records in cloud storage",
        "Keep records for at least 7 years (IRS statute of limitations)",
    ]

    for tip in tips:
        ws.cell(row=row, column=1, value="\u2022").font = DATA_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=tip).font = DATA_FONT
        ws.cell(row=row, column=2).alignment = WRAP
        ws.row_dimensions[row].height = 20
        row += 1

    # References
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value="IRS References").font = SUBTITLE_FONT
    ws.row_dimensions[row].height = 28
    row += 1

    refs = [
        "IRS Schedule C (Form 1040): Profit or Loss from Business",
        "IRS Publication 535: Business Expenses",
        "IRS Publication 463: Travel, Gift, and Car Expenses",
        "IRS Publication 587: Business Use of Your Home",
        "Visit: www.irs.gov/businesses/small-businesses-self-employed",
    ]

    for ref in refs:
        ws.cell(row=row, column=1, value="\u2192").font = DATA_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=ref).font = DATA_FONT
        ws.row_dimensions[row].height = 20
        row += 1

    # Disclaimer
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1,
            value="DISCLAIMER: This guide is for informational purposes only and does not "
                  "constitute tax advice. Always consult a qualified tax professional for "
                  "guidance specific to your situation.").font = NOTE_FONT
    ws.cell(row=row, column=1).alignment = WRAP
    ws.row_dimensions[row].height = 32

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    setup_print(ws, 2, row, '1:1', 'portrait')
    return ws


# ============================================================
# TAB 5: DASHBOARD (built last — references other tabs)
# ============================================================

def build_dashboard(wb, theme, theme_name):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = theme["primary"]

    for col, w in enumerate([3, 22, 18, 18, 18, 18, 3], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title bar
    ws.merge_cells('A1:G1')
    title = ws.cell(row=1, column=1, value="TAX DEDUCTION TRACKER")
    title.font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
    title.fill = header_fill(theme)
    title.alignment = CENTER
    ws.row_dimensions[1].height = 48

    ws.merge_cells('A2:G2')
    sub = ws.cell(row=2, column=1, value="Small Business & Self-Employed Tax Write-Off Organizer")
    sub.font = Font(name='Arial', size=11, color=theme["muted_text"])
    sub.fill = light_fill(theme)
    sub.alignment = CENTER
    ws.row_dimensions[2].height = 28

    # Tax year selector
    ws.cell(row=4, column=2, value="Tax Year:").font = LABEL_FONT
    ws.cell(row=4, column=2).alignment = RIGHT
    year_cell = ws.cell(row=4, column=3, value=2025)
    year_cell.font = Font(name='Arial', size=14, bold=True, color='0000FF')
    year_cell.alignment = CENTER
    year_cell.protection = Protection(locked=False)
    year_cell.border = Border(bottom=Side(style='medium', color=theme["primary"]))

    dv_year = DataValidation(
        type='list', formula1='"2025,2026,2027"',
        allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(dv_year)
    dv_year.add(year_cell)

    ws.row_dimensions[4].height = 30

    log_range_end = 501
    year_ref = '$C$4'

    # --- KPI Cards ---
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 36

    # Card 1: Total Deductions (deductible amount, with meals 50% applied)
    ws.cell(row=6, column=2, value="Total Deductible").font = LABEL_FONT
    ws.cell(row=6, column=2).alignment = CENTER
    cell = ws.cell(row=7, column=2)
    cell.value = f'=\'Category Summary\'!E{3 + NUM_CATS + 1}'
    cell.font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # Card 2: Total Entries
    ws.cell(row=6, column=3, value="Total Entries").font = LABEL_FONT
    ws.cell(row=6, column=3).alignment = CENTER
    cell = ws.cell(row=7, column=3)
    cell.value = (f'=SUMPRODUCT((YEAR(\'Deduction Log\'!A2:A{log_range_end})={year_ref})'
                  f'*(\'Deduction Log\'!A2:A{log_range_end}<>""))')
    cell.font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    cell.number_format = '#,##0'
    cell.alignment = CENTER

    # Card 3: Average Deduction
    ws.cell(row=6, column=4, value="Avg. Deduction").font = LABEL_FONT
    ws.cell(row=6, column=4).alignment = CENTER
    cell = ws.cell(row=7, column=4)
    cell.value = (f'=IF(C7=0,0,SUMPRODUCT((YEAR(\'Deduction Log\'!A2:A{log_range_end})={year_ref})'
                  f'*\'Deduction Log\'!E2:E{log_range_end})/C7)')
    cell.font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # Card 4: Largest Deduction
    ws.cell(row=6, column=5, value="Largest Single").font = LABEL_FONT
    ws.cell(row=6, column=5).alignment = CENTER
    cell = ws.cell(row=7, column=5)
    cell.value = (f'=MAXIFS(\'Deduction Log\'!E2:E{log_range_end},'
                  f'YEAR(\'Deduction Log\'!A2:A{log_range_end}),{year_ref})')
    cell.font = Font(name='Arial', size=18, bold=True, color=theme["primary"])
    cell.number_format = '"$"#,##0.00'
    cell.alignment = CENTER

    # --- Deductions by Category (top section) ---
    ws.row_dimensions[9].height = 28
    ws.merge_cells('B9:F9')
    ws.cell(row=9, column=2, value="Deductions by Category").font = SUBTITLE_FONT

    # Chart data: category + deductible amount from Category Summary
    chart_data_start = 10
    ws.cell(row=chart_data_start, column=2, value="Category").font = SMALL_FONT
    ws.cell(row=chart_data_start, column=3, value="Amount").font = SMALL_FONT

    for i, cat in enumerate(CATEGORIES):
        row = chart_data_start + 1 + i
        ws.cell(row=row, column=2, value=cat).font = SMALL_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=3).value = f'=\'Category Summary\'!E{4 + i}'
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'

    # Pie chart
    pie = PieChart()
    pie.title = None
    pie.style = 10
    pie.width = 18
    pie.height = 12

    data = Reference(ws, min_col=3, min_row=chart_data_start,
                     max_row=chart_data_start + NUM_CATS)
    cats = Reference(ws, min_col=2, min_row=chart_data_start + 1,
                     max_row=chart_data_start + NUM_CATS)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)

    # Color the slices with theme palette
    for i in range(min(NUM_CATS, len(theme["chart_palette"]))):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = theme["chart_palette"][i]
        pie.series[0].data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = False
    pie.dataLabels.showVal = False

    ws.add_chart(pie, "D10")

    # --- Monthly Deduction Trend ---
    monthly_section_row = chart_data_start + NUM_CATS + 3
    ws.merge_cells(f'B{monthly_section_row}:F{monthly_section_row}')
    ws.cell(row=monthly_section_row, column=2, value="Monthly Deduction Trend").font = SUBTITLE_FONT
    ws.row_dimensions[monthly_section_row].height = 28

    md_start = monthly_section_row + 1
    ws.cell(row=md_start, column=2, value="Month").font = SMALL_FONT
    ws.cell(row=md_start, column=3, value="Amount").font = SMALL_FONT

    for i, month in enumerate(MONTHS_SHORT):
        row = md_start + 1 + i
        month_num = i + 1
        ws.cell(row=row, column=2, value=month).font = SMALL_FONT
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3).value = (
            f'=SUMPRODUCT((YEAR(\'Deduction Log\'!A2:A{log_range_end})={year_ref})'
            f'*(MONTH(\'Deduction Log\'!A2:A{log_range_end})={month_num})'
            f'*\'Deduction Log\'!E2:E{log_range_end})')
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'

    # Bar chart for monthly trend
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = None
    bar.y_axis.title = "Amount ($)"
    bar.y_axis.numFmt = '"$"#,##0'
    bar.x_axis.title = None
    bar.width = 28
    bar.height = 12

    data = Reference(ws, min_col=3, min_row=md_start, max_row=md_start + 12)
    cats_ref = Reference(ws, min_col=2, min_row=md_start + 1, max_row=md_start + 12)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.shape = 4
    bar.series[0].graphicalProperties.solidFill = theme["primary"]

    chart_anchor_row = md_start + 14
    ws.add_chart(bar, f"B{chart_anchor_row}")

    # --- Top 5 Largest Deductions ---
    top5_row = chart_anchor_row + 16
    ws.merge_cells(f'B{top5_row}:F{top5_row}')
    ws.cell(row=top5_row, column=2, value="Top 5 Largest Deductions").font = SUBTITLE_FONT
    ws.row_dimensions[top5_row].height = 28

    t5_headers = ["Rank", "Amount", "Description", "Category"]
    style_header_row(ws, top5_row + 1, 2, 5, t5_headers, theme)

    for i in range(5):
        row = top5_row + 2 + i
        style_data_row(ws, row, 2, 5, i)
        ws.cell(row=row, column=2, value=f'#{i + 1}').font = DATA_FONT_BOLD
        ws.cell(row=row, column=2).alignment = CENTER

        # LARGE to get nth largest
        ws.cell(row=row, column=3).value = (
            f'=IFERROR(LARGE(\'Deduction Log\'!E$2:E${log_range_end},{i + 1}),"")')
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=3).alignment = RIGHT

        # INDEX/MATCH to get description of nth largest
        ws.cell(row=row, column=4).value = (
            f'=IFERROR(INDEX(\'Deduction Log\'!B$2:B${log_range_end},'
            f'MATCH(LARGE(\'Deduction Log\'!E$2:E${log_range_end},{i + 1}),'
            f'\'Deduction Log\'!E$2:E${log_range_end},0)),"")' )
        ws.cell(row=row, column=4).alignment = LEFT

        # INDEX/MATCH to get category
        ws.cell(row=row, column=5).value = (
            f'=IFERROR(INDEX(\'Deduction Log\'!C$2:C${log_range_end},'
            f'MATCH(LARGE(\'Deduction Log\'!E$2:E${log_range_end},{i + 1}),'
            f'\'Deduction Log\'!E$2:E${log_range_end},0)),"")' )
        ws.cell(row=row, column=5).alignment = LEFT

    # Protection
    ws.protection.sheet = True
    ws.protection.enable()

    return ws


# ============================================================
# MAIN GENERATOR
# ============================================================

def generate(theme_name, theme):
    wb = Workbook()
    wb.remove(wb.active)

    # Build tabs in order
    build_deduction_log(wb, theme)
    build_category_summary(wb, theme)
    build_annual_summary(wb, theme)
    build_deduction_guide(wb, theme)
    build_dashboard(wb, theme, theme_name)

    # Reorder: Dashboard first
    wb.move_sheet("Dashboard", offset=-4)

    # Set Dashboard as active
    wb.active = wb.sheetnames.index("Dashboard")

    # Save
    safe_name = theme_name.replace(" ", "-")
    filename = f"Tax-Deduction-Tracker-{safe_name}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"  Generated: {filepath}")
    return filepath


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Generating #010 Tax Deduction Tracker...")
    print(f"Output: {OUTPUT_DIR}")
    print()

    files = []
    for theme_name, theme in THEMES.items():
        f = generate(theme_name, theme)
        files.append(f)

    print()
    print(f"Done! {len(files)} files generated.")
    for f in files:
        print(f"  {os.path.basename(f)}")


if __name__ == "__main__":
    main()
